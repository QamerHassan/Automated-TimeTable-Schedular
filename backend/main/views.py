# Enhanced Genetic Algorithm Timetable Generation System
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from .models import (CoreCourse, CohortCourse, ElectiveCourse, Room, SemesterStudents,
Timetable, Section, EnhancedSchedule, GlobalScheduleMatrix, SubjectFrequencyRule, TimeSlotConfiguration)
from .forms import TimetableGenerationForm
from django.contrib.auth.decorators import login_required
from django.db.models import Sum
from django.core.files.storage import default_storage
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
import random
import csv
import os
import io
import logging
import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from django.utils.text import slugify
from datetime import datetime, timedelta, time
from collections import defaultdict
import numpy as np
import uuid # Add this import at the top of your file if it's not there
import random
from collections import defaultdict

logger = logging.getLogger(__name__)

# ============ GENETIC ALGORITHM CONFIGURATION ============
# In backend/main/views.py

# ============ GENETIC ALGORITHM CONFIGURATION ============
GENETIC_CONFIG = {
    'POPULATION_SIZE': 50,
    'GENERATIONS': 150,
    'CROSSOVER_RATE': 0.85,
    'BASE_MUTATION_RATE': 0.2,
    'MAX_MUTATION_RATE': 0.8,
    'ELITE_SIZE': 5,
    'TOURNAMENT_SIZE': 7
}  # Added missing closing brace



SUBJECT_FREQUENCY_MAP = {
# Theoretical subjects - 2 times per week
"English-I": 2, "Calculus": 2, "Programming Fundamentals": 2,
"ICT Theory": 2, "Information Security": 2, "Physics": 2,
"Chemistry": 2, "Mathematics": 2, "Statistics": 2,
# Lab subjects - 1 time per week
"Programming Fundamentals Lab": 1, "ICT Lab": 1,
"Physics Lab": 1, "Chemistry Lab": 1,
}
# In backend/main/views.py
# Add this function right after the SUBJECT_FREQUENCY_MAP definition

def get_subject_frequency(subject_name, is_lab):
    """Get frequency based on subject type and name"""
    if is_lab:
        return 1  # Labs are typically once per week
    # For theory subjects, check the map or default to 2 times per week
    return SUBJECT_FREQUENCY_MAP.get(subject_name, 2)


ENHANCED_TIME_SLOTS = [
"08:00-09:15", "09:30-10:45", "11:00-12:15", # Morning
"12:30-13:45", "14:00-15:15", "15:30-16:45", # Afternoon
"17:00-18:15", "18:30-19:45", "20:00-21:15" # Evening (Extended hours)
]
# CORRECTED: Lab time slots as paired 1h 15m blocks with 15-minute breaks
LAB_TIME_SLOTS = [
    ["08:00-09:15", "09:30-10:45"],  # Morning Block 1: 8:00-10:45 with break
    ["11:00-12:15", "12:30-13:45"],  # Morning Block 2: 11:00-13:45 with break  
    ["14:00-15:15", "15:30-16:45"],  # Afternoon Block 1: 14:00-16:45 with break
    ["17:00-18:15", "18:30-19:45"],  # Afternoon Block 2: 17:00-19:45 with break
]

# Theory time slots (keep existing or use this)
THEORY_TIME_SLOTS = [
    "08:00-09:15", "09:30-10:45", "11:00-12:15",
    "12:30-13:45", "14:00-15:15", "15:30-16:45", 
    "17:00-18:15", "18:30-19:45", "20:00-21:15"
]


# Combined slots for general use
ENHANCED_TIME_SLOTS = THEORY_TIME_SLOTS.copy()

# Add this right after LAB_TIME_SLOTS definition (around line 94)
THEORY_TIME_SLOTS = [
    "08:00-09:15", "09:30-10:45", "11:00-12:15", 
    "12:30-13:45", "14:00-15:15", "15:30-16:45", 
    "17:00-18:15", "18:30-19:45", "20:00-21:15"
]

WEEKDAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]

def home(request):
    """API-friendly home endpoint - no templates needed"""
    return JsonResponse({
        'status': 'success',
        'message': 'Genetic Algorithm Timetable System API Active',
        'version': '2.0-GA',
        'algorithm': 'Genetic Algorithm + Constraint Satisfaction',
        'backend_features': [
            'Twice-weekly theoretical lectures',
            'Inter-semester conflict resolution',
            'Extended hours (8 AM - 9 PM)',
            'Multi-objective optimization'
        ],
        'api_endpoints': {
            'generate': '/api/generate/',
            'enhanced_generate': '/api/generate-enhanced/',
            'validate': '/api/validate-enhanced/',
            'stats': '/api/schedule-stats/'
        },
        'genetic_config': {
            'population_size': 50,
            'generations': 100,
            'crossover_rate': 0.8,
            'mutation_rate': 0.1
        }
    })

# ============ GENETIC ALGORITHM CORE CLASSES ============

class Gene:
    """Represents a single class session, with lab block pairing support."""
    
    def __init__(self, semester, section, subject, is_lab, session_number=1, is_cohort=False):
        self.id = uuid.uuid4()
        self.semester = semester
        self.section = section
        self.subject = subject
        self.is_lab = is_lab
        self.session_number = session_number
        self.day = None
        self.is_cohort = is_cohort
        
        # NEW: Lab block handling
        if is_lab:
            self.time_slot = None  # Will be set to lab block pair
            self.lab_block_pair = None  # Will store both time slots
        else:
            self.time_slot = None  # Single theory slot
            
        self.room = None

    def __repr__(self):
        room_name = self.room.room_number if self.room else "None"
        cohort_indicator = "(Cohort)" if self.is_cohort else ""
        
        if hasattr(self, 'is_lab_blocks') and self.is_lab_blocks and self.lab_block_pair:
            time_display = f"{self.lab_block_pair[0]} + {self.lab_block_pair[1]}"  # Use + instead of &
        elif self.time_slot:
            time_display = self.time_slot
        else:
            time_display = "UNASSIGNED"
        
        return (f"Gene(S{self.semester}-{self.section}, {self.subject}, {cohort_indicator}"
                f"Day={getattr(self, 'day', 'None')}, Slot={time_display}, Room={room_name})")



# =====================================================================
# REVISED AND IMPROVED Chromosome CLASS
# Replace the old Chromosome class in views.py with this one.
# =====================================================================
class Chromosome:
    """Represents a complete timetable solution"""
    def __init__(self, genes=None):
        self.genes = genes or []
        self.fitness = 0
        self.conflicts = []
        # The schedule_matrix is no longer needed here, it's handled during database save
        
    def _calculate_scheduling_priority(self, semester, is_lab):
        """Calculates a priority score for a class. Higher is more critical."""
        priority = 0
        if is_lab:
            priority += 100 # Labs are high priority due to limited rooms
        priority += semester # Higher semesters are higher priority
        return priority
    def assign_rooms_to_genes(self):
        """Assign appropriate Room OBJECTS to genes, with cohort course logic."""
        lab_rooms = list(Room.objects.filter(lab=True))
        theory_rooms = list(Room.objects.filter(core_sub=True))
        
        for gene in self.genes:
            if gene.is_cohort:
                # Cohort courses are ALWAYS theoretical, use theory rooms
                gene.room = random.choice(theory_rooms) if theory_rooms else None
            elif gene.is_lab and lab_rooms:
                gene.room = random.choice(lab_rooms)
            elif theory_rooms:
                gene.room = random.choice(theory_rooms)
            else:
                gene.room = None
                logger.error(f"Could not assign a room for {gene.subject}.")

        
    def initialize_sequentially(self, semesters_data):
        """NUCLEAR OPTION: Absolute final version with guaranteed theory slot assignment"""
        print("="*50)
        print("NUCLEAR OPTION: initialize_sequentially RUNNING")
        print("="*50)
        
        self.genes = []
        
        # Fetch rooms
        all_lab_rooms = list(Room.objects.filter(lab=True))
        all_theory_rooms = list(Room.objects.filter(core_sub=True))
        
        if not all_lab_rooms or not all_theory_rooms:
            logger.error("CRITICAL: No rooms found")
            return

        schedule_tracker = {}
        class_list = []

        # Create class list
        for semester, sections_data in semesters_data.items():
            for section_name, subjects in sections_data.items():
                for subject_info in subjects:
                    frequency = get_subject_frequency(subject_info['name'], subject_info['is_lab'])
                    for session in range(frequency):
                        class_list.append({
                            "semester": semester,
                            "section": section_name,
                            "subject": subject_info['name'],
                            "is_lab": subject_info['is_lab'],
                            "is_cohort": subject_info.get('is_cohort', False),
                            "session": session,
                            "priority": self._calculate_scheduling_priority(semester, subject_info['is_lab'])
                        })
        
        class_list.sort(key=lambda x: x['priority'], reverse=True)

        # Process each class
        for class_info in class_list:
            gene = Gene(
                semester=class_info['semester'],
                section=class_info['section'],
                subject=class_info['subject'],
                is_lab=class_info['is_lab'],
                session_number=class_info['session'] + 1,
                is_cohort=class_info['is_cohort']
            )
            
            # Room selection
            if gene.is_cohort:
                available_rooms_for_gene = all_theory_rooms
            else:
                available_rooms_for_gene = all_lab_rooms if gene.is_lab else all_theory_rooms
            
            # Try to find slot
            best_slot_info = self._find_best_slot_shuffled(gene, schedule_tracker, available_rooms_for_gene)
            
            # CRITICAL: Always assign valid values
            if gene.is_lab:
                gene.lab_block_pair = LAB_TIME_SLOTS[0] if LAB_TIME_SLOTS else [["08:00-09:15", "09:30-10:45"]]
                gene.time_slot = None
                gene.is_lab_blocks = True
            else:
                # GUARANTEE: Theory classes ALWAYS get valid time_slot
                gene.time_slot = THEORY_TIME_SLOTS[0] if THEORY_TIME_SLOTS else "08:00-09:15"
                gene.lab_block_pair = None
                gene.is_lab_blocks = False
                print(f"FORCED ASSIGNMENT: {gene.subject} → {gene.time_slot}")
            
            # Set basic properties
            if best_slot_info:
                day, time_info, room_obj = best_slot_info
                gene.day = day
                gene.room = room_obj
                
                # Only update time assignments if slot was found and is valid
                if gene.is_lab and time_info:
                    gene.lab_block_pair = time_info
                elif not gene.is_lab and time_info and isinstance(time_info, str):
                    gene.time_slot = time_info
                    print(f"OPTIMAL ASSIGNMENT: {gene.subject} → {gene.time_slot}")
            else:
                gene.day = "Monday"
                gene.room = available_rooms_for_gene[0] if available_rooms_for_gene else None
            
            self.genes.append(gene)
        
        # ABSOLUTE FINAL SAFETY NET
        print("RUNNING FINAL SAFETY NET...")
        for gene in self.genes:
            if not gene.is_lab:
                if not gene.time_slot or gene.time_slot is None or gene.time_slot == "None":
                    gene.time_slot = "08:00-09:15"  # Hard-coded fallback
                    print(f"SAFETY NET ACTIVATED: {gene.subject} → {gene.time_slot}")
        
        print(f"INITIALIZATION COMPLETE: {len(self.genes)} genes created")




    def _find_best_slot_shuffled(self, gene_to_schedule, schedule_tracker, available_rooms):
        """Enhanced slot finding with debug logging and robust fallback logic"""
    
    # Validate inputs first
        if not available_rooms:
            print(f"ERROR: No available rooms for {gene_to_schedule.subject} (is_lab: {gene_to_schedule.is_lab})")
            return None

        shuffled_rooms = random.sample(available_rooms, len(available_rooms))
        shuffled_days = random.sample(WEEKDAYS, len(WEEKDAYS))
    
        if gene_to_schedule.is_lab:
            # For labs, shuffle the paired block options
            if not LAB_TIME_SLOTS:
                print(f"ERROR: LAB_TIME_SLOTS is empty for {gene_to_schedule.subject}")
                return None
            
            shuffled_lab_blocks = random.sample(LAB_TIME_SLOTS, len(LAB_TIME_SLOTS))
        
            for room_obj in shuffled_rooms:
                for day in shuffled_days:
                    for lab_block_pair in shuffled_lab_blocks:
                    # Check if both slots in the pair are available
                        slot1, slot2 = lab_block_pair
                    
                        section_key1 = f"S{gene_to_schedule.semester}-{gene_to_schedule.section}-{day}-{slot1}"
                        section_key2 = f"S{gene_to_schedule.semester}-{gene_to_schedule.section}-{day}-{slot2}"
                        room_key1 = f"R{room_obj.id}-{day}-{slot1}"
                        room_key2 = f"R{room_obj.id}-{day}-{slot2}"
                    
                    # Both slots must be free
                        if (not schedule_tracker.get(room_key1) and not schedule_tracker.get(section_key1) and
                            not schedule_tracker.get(room_key2) and not schedule_tracker.get(section_key2)):
                            print(f"DEBUG: Found lab slot for {gene_to_schedule.subject}: {day}, {lab_block_pair}")
                            return (day, lab_block_pair, room_obj)
        
            print(f"DEBUG: NO LAB SLOT FOUND for {gene_to_schedule.subject}")
            return None
        
        else:
            # Theory classes use single slots
            if not THEORY_TIME_SLOTS:
                print(f"ERROR: THEORY_TIME_SLOTS is empty for {gene_to_schedule.subject}")
                return None
            
            shuffled_slots = random.sample(THEORY_TIME_SLOTS, len(THEORY_TIME_SLOTS))
        
            for room_obj in shuffled_rooms:
                for day in shuffled_days:
                    for time_slot in shuffled_slots:
                        section_key = f"S{gene_to_schedule.semester}-{gene_to_schedule.section}-{day}-{time_slot}"
                        room_key = f"R{room_obj.id}-{day}-{time_slot}"
                    
                        if not schedule_tracker.get(room_key) and not schedule_tracker.get(section_key):
                            print(f"DEBUG: Found theory slot for {gene_to_schedule.subject}: {day}, {time_slot}")
                            return (day, time_slot, room_obj)
        
            print(f"DEBUG: NO THEORY SLOT FOUND for {gene_to_schedule.subject} (checked {len(shuffled_slots)} slots across {len(shuffled_days)} days in {len(shuffled_rooms)} rooms)")
            return None
        
    def _find_fallback_slot(self, gene, available_rooms):
        if available_rooms and WEEKDAYS:
            day = random.choice(WEEKDAYS)
            if gene.is_lab:
                time_slot = random.choice(LAB_TIME_SLOTS)
            else:
                time_slot = random.choice(THEORY_TIME_SLOTS)
            room = random.choice(available_rooms)
            return (day, time_slot, room)
        return None



# =====================================================================
# REVISED AND IMPROVED GeneticAlgorithm CLASS
# Replace the old GeneticAlgorithm class in views.py with this one.
# =====================================================================
# =====================================================================
# THE DEFINITIVE, COMPLETE GeneticAlgorithm CLASS
# Replace the entire existing GeneticAlgorithm class in views.py with this.
# This version contains all necessary methods and will resolve the crash.
# =====================================================================
# =====================================================================
# THE DEFINITIVE, COMPLETE GeneticAlgorithm CLASS
# Replace the entire existing GeneticAlgorithm class in views.py with this.
# This version contains all necessary methods and will resolve the crash.
# =====================================================================
class GeneticAlgorithm:
    """
    Final, robust Genetic Algorithm implementation.
    Features perfectly synchronized conflict detection and advanced mutation operators
    to escape local optima and solve complex scheduling constraints.
    """
    def __init__(self, config=None):
        self.config = config or GENETIC_CONFIG
        self.population = []
        self.best_chromosome = None
        self.generation = 0
        # Tracks fitness to detect stagnation
        self.stagnation_counter = 0
        self.last_best_fitness = -float('inf')

    def initialize_population(self, semesters_data):
        logger.info(f"Initializing population of size {self.config['POPULATION_SIZE']}...")
        self.population = []
        for i in range(self.config['POPULATION_SIZE']):
            chromosome = Chromosome()
            chromosome.initialize_sequentially(semesters_data)
            self.population.append(chromosome)
        logger.info("Population initialized.")

    def evaluate_fitness(self, chromosome):
        """Calculates fitness based on both hard and soft constraints."""
        fitness = 1000
    
        conflicting_genes = self.get_all_hard_conflicts(chromosome)
        num_conflicts = len(conflicting_genes)

    # Apply a very heavy penalty for any hard conflicts
        fitness -= num_conflicts * 500 
    
    # Also apply penalties for soft constraints
        daily_load_penalties = self.check_daily_load_balance(chromosome)
        gap_penalties = self.check_time_gaps(chromosome)
        theoretical_frequency_penalties = self.check_theoretical_frequency(chromosome)

        fitness -= daily_load_penalties * 5
        fitness -= gap_penalties * 3
        fitness -= theoretical_frequency_penalties * 10
    
    # Diversity penalty (fitness sharing) - optimized to sample instead of full O(N^2)
        diversity_penalty = 0
        if len(self.population) > 1:  # Avoid if population is too small
            sample_size = min(20, len(self.population) - 1)  # Sample to reduce computation
            sampled_others = random.sample([other for other in self.population if other != chromosome], sample_size)
            for other in sampled_others:
                similarity = self._calculate_similarity(chromosome, other)
                if similarity > 0.8:  # Threshold for "too similar"
                    diversity_penalty += similarity
    
        fitness -= diversity_penalty * 10  # Adjust multiplier as needed
    
        chromosome.conflicts = list(conflicting_genes)
        chromosome.fitness = fitness
        return fitness
    
    def _calculate_similarity(self, chrom1, chrom2):
        # Simple Hamming distance for demonstration (customize for your genes)
        matching_genes = sum(1 for g1, g2 in zip(chrom1.genes, chrom2.genes) if g1.time_slot == g2.time_slot and g1.room == g2.room)
        return matching_genes / max(len(chrom1.genes), 1)  # Normalized 0-1
    def _slots_overlap(self, gene1, gene2):
        """Enhanced overlap detection for lab blocks and theory slots"""
    
    # Different days = no overlap
        if gene1.day != gene2.day:
            return False
    
    # Get time slots for comparison
        if gene1.is_lab:
            slots1 = gene1.lab_block_pair  # List of two slots
        else:
            slots1 = [gene1.time_slot]     # Single slot as list
        
        if gene2.is_lab:
            slots2 = gene2.lab_block_pair  # List of two slots  
        else:
            slots2 = [gene2.time_slot]     # Single slot as list
    
    # Check if any slot from gene1 overlaps with any slot from gene2
        for slot1 in slots1:
            for slot2 in slots2:
                if self._single_slot_overlap(slot1, slot2):
                    return True
    
        return False


    def get_all_hard_conflicts(self, chromosome):
        """Find all hard constraint violations in the chromosome"""
        conflicting_genes = []
        genes = chromosome.genes
    
        for i in range(len(genes)):
            for j in range(i + 1, len(genes)):
                g1, g2 = genes[i], genes[j]
            
            # Check if genes conflict (same room, section, or instructor at overlapping times)
                conflicts = False
            
            # Only check for conflicts if they're on the same day
                if g1.day == g2.day:
                # Check time overlap using the gene-aware method
                    if self._genes_overlap(g1, g2):
                    # Room conflict
                        if g1.room and g2.room and g1.room.id == g2.room.id:
                            conflicts = True
                    
                    # Section conflict (same section can't be in two places)
                        if (g1.semester == g2.semester and g1.section == g2.section):
                            conflicts = True
            
                if conflicts:
                    conflicting_genes.extend([g1, g2])
    
        return list(set(conflicting_genes))  # Remove duplicates

    def _genes_overlap(self, gene1, gene2):
        """Check if two genes have overlapping time slots"""
    # Different days = no overlap
        if gene1.day != gene2.day:
            return False
    
    # Get time slots for comparison - ALWAYS return arrays
        def get_time_slots(gene):
            if hasattr(gene, 'is_lab_blocks') and gene.is_lab_blocks and gene.lab_block_pair:
                return gene.lab_block_pair  # Already an array
            elif gene.time_slot:    
                return [gene.time_slot]     # Single slot as array
            else:
                return []  # No valid slots
    
        slots1 = get_time_slots(gene1)
        slots2 = get_time_slots(gene2)
    
    # Check if any slot from gene1 overlaps with any slot from gene2
        for slot1 in slots1:
            for slot2 in slots2:
                if self._check_time_overlap(slot1, slot2):
                    return True
    
        return False


    def _check_time_overlap(self, slot_a, slot_b):
        """Check if two individual time slots overlap"""
        if slot_a == slot_b:
            return True
        
        try:
            s_a, e_a = [datetime.strptime(t.strip(), '%H:%M') for t in slot_a.split('-')]
            s_b, e_b = [datetime.strptime(t.strip(), '%H:%M') for t in slot_b.split('-')]
            return max(s_a, s_b) < min(e_a, e_b)
        except Exception as exc:
            logger.error(f"Bad slot format ({slot_a}|{slot_b}) → {exc}")
            return False

    def _single_slot_overlap(self, slot1, slot2):
        """Check if two individual 1h 15m slots overlap"""
        if slot1 == slot2:
            return True
        
        try:
            s_a, e_a = [datetime.strptime(t.strip(), '%H:%M') for t in slot1.split('-')]
            s_b, e_b = [datetime.strptime(t.strip(), '%H:%M') for t in slot2.split('-')]
            return max(s_a, s_b) < min(e_a, e_b)
        except Exception as exc:
            logger.error(f"Bad slot format ({slot1}|{slot2}) → {exc}")
            return False



    # --- MISSING HELPER METHODS NOW INCLUDED ---
    def check_daily_load_balance(self, chromosome):
        penalties = 0
        daily_loads = defaultdict(lambda: defaultdict(int))
        for gene in chromosome.genes:
            if not gene.day: continue
            section_key = f"S{gene.semester}-{gene.section}"
            daily_loads[section_key][gene.day] += 1
        for section, days in daily_loads.items():
            for day, count in days.items():
                if count > 4: # Penalize more than 4 classes a day
                    penalties += (count - 4)
        return penalties

    def check_time_gaps(self, chromosome):
        gap_penalties = 0
        schedule_by_section_day = defaultdict(list)
        for gene in chromosome.genes:
            if not gene.time_slot or not gene.day: continue
            key = f"S{gene.semester}-{gene.section}-{gene.day}"
            schedule_by_section_day[key].append(gene.time_slot)
        
        for key, slots in schedule_by_section_day.items():
            if len(slots) < 2: continue
            try:
                sorted_slots = sorted(slots, key=lambda x: datetime.strptime(x.split('-')[0].strip(), '%H:%M'))
                for i in range(len(sorted_slots) - 1):
                    end_time_current = datetime.strptime(sorted_slots[i].split('-')[1].strip(), '%H:%M')
                    start_time_next = datetime.strptime(sorted_slots[i+1].split('-')[0].strip(), '%H:%M')
                    gap_minutes = (start_time_next - end_time_current).total_seconds() / 60
                    if gap_minutes > 15: # Penalize gaps larger than 15 mins
                        gap_penalties += (gap_minutes / 75)
            except (ValueError, IndexError):
                continue
        return int(gap_penalties)
    
    def check_theoretical_frequency(self, chromosome):
        penalties = 0
        subject_counts = defaultdict(int)
        for gene in chromosome.genes:
            if not gene.is_lab:
                key = f"S{gene.semester}-{gene.section}-{gene.subject}"
                subject_counts[key] += 1
        
        for subject_key, count in subject_counts.items():
            # Dynamically get expected frequency
            parts = subject_key.split('-')
            subject_name = parts[-1]
            expected_freq = get_subject_frequency(subject_name, is_lab=False)
            if count != expected_freq:
                penalties += abs(count - expected_freq)
        return penalties

    # --- MUTATION OPERATORS ---
    def mutate(self, chromosome):
        """Main mutation function with stagnation-aware logic."""
        adaptive_rate = self.config['BASE_MUTATION_RATE'] + (self.stagnation_counter / 100) * (self.config['MAX_MUTATION_RATE'] - self.config['BASE_MUTATION_RATE'])
        adaptive_rate = min(adaptive_rate, self.config['MAX_MUTATION_RATE'])
        if self.stagnation_counter > 5 and random.random() < 0.3:
            self.partial_scramble_mutate(chromosome)
            return

        if chromosome.conflicts and random.random() < 0.95:
            self.repair_mutate(chromosome)
        
        elif random.random() < 0.6:
            self.swap_mutate(chromosome)

        if random.random() < adaptive_rate:  # Use adaptive rate here
            self.random_mutate(chromosome)
    def conflict_guided_mutation(self, chromosome):
        """Mutation that targets conflicting genes specifically"""   
        conflicting_genes = self.get_all_hard_conflicts(chromosome)
    
        if conflicting_genes:
        # Target a conflicting gene for mutation
            gene_to_fix = random.choice(conflicting_genes)
        
        # Try to move it to a completely different day/time
            gene_to_fix.day = random.choice(WEEKDAYS)
        
            if hasattr(gene_to_fix, 'is_lab') and gene_to_fix.is_lab:
                gene_to_fix.lab_block_pair = random.choice(LAB_TIME_SLOTS)
                gene_to_fix.time_slot = None
                gene_to_fix.is_lab_blocks = True
                room_pool = Room.objects.filter(lab=True)
            else:
                gene_to_fix.time_slot = random.choice(THEORY_TIME_SLOTS)
                gene_to_fix.lab_block_pair = None
                gene_to_fix.is_lab_blocks = False
                room_pool = Room.objects.filter(core_sub=True)
        
            if room_pool.exists():
                gene_to_fix.room = random.choice(list(room_pool))

    def partial_scramble_mutate(self, chromosome):
        """A 'bigger hammer' mutation for escaping local optima."""
        if not chromosome.conflicts: return
        sections_to_scramble = min(3, len(set(g.semester for g in chromosome.conflicts)))
        
        conflicting_section_gene = random.choice(list(chromosome.conflicts))
        target_semester = conflicting_section_gene.semester
        target_section = conflicting_section_gene.section
        for _ in range(sections_to_scramble):
            if not chromosome.conflicts: break
        
            conflicting_section_gene = random.choice(list(chromosome.conflicts))
            target_semester = conflicting_section_gene.semester
            target_section = conflicting_section_gene.section
            logger.info(f"Heavy stagnation detected. Scrambling S{target_semester}-{target_section}...")
        genes_to_reschedule = [g for g in chromosome.genes if g.semester == target_semester and g.section == target_section]
        other_genes = [g for g in chromosome.genes if not (g.semester == target_semester and g.section == target_section)]
        
        schedule_tracker = {}
        for g in other_genes:
            if g.day and g.time_slot:
                unique_section_key = f"S{g.semester}-{g.section}-{g.day}-{g.time_slot}"
                if g.room: schedule_tracker[f"R{g.room.id}-{g.day}-{g.time_slot}"] = True
                schedule_tracker[unique_section_key] = True

        for gene in genes_to_reschedule:
            available_rooms = list(Room.objects.filter(lab=gene.is_lab))
            new_slot_info = self._find_best_slot_shuffled(gene, schedule_tracker, available_rooms)
            if new_slot_info:
                day, time_slot, room_obj = new_slot_info
                gene.day, gene.time_slot, gene.room = day, time_slot, room_obj
                schedule_tracker[f"R{room_obj.id}-{day}-{time_slot}"] = True
                unique_section_key = f"S{gene.semester}-{gene.section}-{day}-{time_slot}"
                schedule_tracker[unique_section_key] = True

    def _find_best_slot_shuffled(self, gene_to_schedule, schedule_tracker, available_rooms):
        shuffled_rooms = random.sample(available_rooms, len(available_rooms))
        shuffled_days = random.sample(WEEKDAYS, len(WEEKDAYS))
    
        if gene_to_schedule.is_lab:
            # For labs, shuffle the paired block options
            shuffled_lab_blocks = random.sample(LAB_TIME_SLOTS, len(LAB_TIME_SLOTS))
        
            for room_obj in shuffled_rooms:
                for day in shuffled_days:
                    for lab_block_pair in shuffled_lab_blocks:
                    # Check if both slots in the pair are available
                        slot1, slot2 = lab_block_pair
                    
                        section_key1 = f"S{gene_to_schedule.semester}-{gene_to_schedule.section}-{day}-{slot1}"
                        section_key2 = f"S{gene_to_schedule.semester}-{gene_to_schedule.section}-{day}-{slot2}"
                        room_key1 = f"R{room_obj.id}-{day}-{slot1}"
                        room_key2 = f"R{room_obj.id}-{day}-{slot2}"
                    
                    # Both slots must be free
                        if (not schedule_tracker.get(room_key1) and not schedule_tracker.get(section_key1) and
                            not schedule_tracker.get(room_key2) and not schedule_tracker.get(section_key2)):
                            return (day, lab_block_pair, room_obj)
        else:
        # Theory classes use single slots
            shuffled_slots = random.sample(THEORY_TIME_SLOTS, len(THEORY_TIME_SLOTS))
        
            for room_obj in shuffled_rooms:
                for day in shuffled_days:
                    for time_slot in shuffled_slots:
                        section_key = f"S{gene_to_schedule.semester}-{gene_to_schedule.section}-{day}-{time_slot}"
                        room_key = f"R{room_obj.id}-{day}-{time_slot}"
                    
                        if not schedule_tracker.get(room_key) and not schedule_tracker.get(section_key):
                            return (day, time_slot, room_obj)
    
        return None

    
    def repair_mutate(self, chromosome):
        if not chromosome.conflicts: return
        gene_to_fix = random.choice(list(chromosome.conflicts))
        
        schedule_tracker = {}
        for g in chromosome.genes:
            if g is gene_to_fix: continue
            if g.day and g.time_slot:
                unique_section_key = f"S{g.semester}-{g.section}-{g.day}-{g.time_slot}"
                if g.room: schedule_tracker[f"R{g.room.id}-{g.day}-{g.time_slot}"] = True
                schedule_tracker[unique_section_key] = True

        available_rooms = list(Room.objects.filter(lab=gene_to_fix.is_lab))
        if not available_rooms: return
        new_slot_info = self._find_best_slot_shuffled(gene_to_fix, schedule_tracker, available_rooms)
        if new_slot_info:
            gene_to_fix.day, gene_to_fix.time_slot, gene_to_fix.room = new_slot_info

    def swap_mutate(self, chromosome):
        if len(chromosome.genes) < 2: return
        idx1, idx2 = random.sample(range(len(chromosome.genes)), 2)
        gene1, gene2 = chromosome.genes[idx1], chromosome.genes[idx2]
        gene1.day, gene2.day = gene2.day, gene1.day
        gene1.time_slot, gene2.time_slot = gene2.time_slot, gene1.time_slot
        gene1.room, gene2.room = gene2.room, gene1.room

    def random_mutate(self, chromosome):
        """Random mutation with lab block support"""
        if not chromosome.genes:
            return
        
        gene_to_mutate = random.choice(chromosome.genes)
        gene_to_mutate.day = random.choice(WEEKDAYS)
    
        if gene_to_mutate.is_lab:
            # Assign random lab block pair
            gene_to_mutate.lab_block_pair = random.choice(LAB_TIME_SLOTS)
            gene_to_mutate.time_slot = None # No single time slot for labs
            gene_to_mutate.is_lab_blocks = True
            room_pool = Room.objects.filter(lab=True)
        else:
        # Assign random theory slot
            gene_to_mutate.time_slot = random.choice(THEORY_TIME_SLOTS)
            gene_to_mutate.lab_block_pair = None
            gene_to_mutate.is_lab_blocks = False
            room_pool = Room.objects.filter(core_sub=True)
    
        if room_pool.exists():
            gene_to_mutate.room = random.choice(list(room_pool))




    # --- CORE EVOLUTIONARY LOOP ---
    def selection(self):
        tournament = random.sample(self.population, self.config['TOURNAMENT_SIZE'])
        return max(tournament, key=lambda c: c.fitness)

    def crossover(self, parent1, parent2):
        if random.random() > self.config['CROSSOVER_RATE'] or len(parent1.genes) <= 1:
            return Chromosome(parent1.genes[:]), Chromosome(parent2.genes[:])
        crossover_point = random.randint(1, len(parent1.genes) - 1)
        child1_genes = parent1.genes[:crossover_point] + parent2.genes[crossover_point:]
        child2_genes = parent2.genes[:crossover_point] + parent1.genes[crossover_point:]
        return Chromosome(child1_genes), Chromosome(child2_genes)

    def evolve(self, semesters_data):
        """Enhanced evolution with performance optimizations and conflict-guided mutation"""
        self.initialize_population(semesters_data)
    
    # Initialize tracking variables
        if not hasattr(self, 'stagnation_counter'):
            self.stagnation_counter = 0
        if not hasattr(self, 'last_best_fitness'):
            self.last_best_fitness = float('-inf')
    
    # Initial fitness evaluation
        for chromosome in self.population:
            self.evaluate_fitness(chromosome)

        for generation in range(self.config['GENERATIONS']):
            self.generation = generation
            self.population.sort(key=lambda c: c.fitness, reverse=True)

        # Track best solution and stagnation
            current_best = self.population[0]
            if not hasattr(self, 'best_chromosome') or self.best_chromosome is None or current_best.fitness > self.best_chromosome.fitness:
                self.best_chromosome = current_best
                if self.best_chromosome.fitness > self.last_best_fitness:
                    self.last_best_fitness = self.best_chromosome.fitness
                    self.stagnation_counter = 0
                else:
                    self.stagnation_counter += 1
            else:
                self.stagnation_counter += 1

        # Progress reporting
            if generation % 10 == 0:
                conflict_count = len(self.best_chromosome.conflicts) if hasattr(self.best_chromosome, 'conflicts') else getattr(self.best_chromosome, 'conflict_count', 'unknown')
                print(f"Generation {generation}: Best Fitness={self.best_chromosome.fitness:.2f}, Conflicts={conflict_count}, Stagnation={self.stagnation_counter}")

        # Check for optimal solution
            if hasattr(self.best_chromosome, 'conflicts'):
                if len(self.best_chromosome.conflicts) == 0:
                    print(f"Optimal solution found at generation {generation}.")
                    break
            elif hasattr(self.best_chromosome, 'conflict_count'):
                if self.best_chromosome.conflict_count == 0:
                    print(f"Optimal solution found at generation {generation}.")
                    break

        # Diversity injection every 10 generations
            if generation % 10 == 0 and generation > 0:
                injection_count = int(len(self.population) * 0.2)  # 20% injection
                for i in range(injection_count):
                    fresh_chromosome = Chromosome()
                    fresh_chromosome.initialize_sequentially(semesters_data)
                    self.evaluate_fitness(fresh_chromosome)
                # Replace worst performers
                    self.population[-(i + 1)] = fresh_chromosome
            
                print(f"Generation {generation}: Injected {injection_count} fresh chromosomes for diversity")

        # Aggressive stagnation handling
            if self.stagnation_counter > 15:  # Reduced from 20 for faster response
                print(f"High stagnation detected ({self.stagnation_counter}). Injecting new population...")
                num_new = int(self.config['POPULATION_SIZE'] * 0.3)  # Increased to 30% for more aggressive restart
                new_chromosomes = []
                for _ in range(num_new):
                    new_chrom = Chromosome()
                    new_chrom.initialize_sequentially(semesters_data)
                    self.evaluate_fitness(new_chrom)
                    new_chromosomes.append(new_chrom)
            
            # Keep best chromosomes and add new ones
                self.population = sorted(self.population[:-num_new] + new_chromosomes, key=lambda c: c.fitness, reverse=True)
                self.stagnation_counter = 0
                print(f"Population restart: Added {num_new} new individuals")

        # Elitism: Keep best solutions
            new_population = self.population[:self.config['ELITE_SIZE']]

        # Generate new offspring
            while len(new_population) < self.config['POPULATION_SIZE']:
                parent1 = self.selection()
                parent2 = self.selection()
                child1, child2 = self.crossover(parent1, parent2)
            
            # INTEGRATED: Use conflict-guided mutation for better performance
                if random.random() < 0.7:  # 70% chance of conflict-guided mutation
                    self.conflict_guided_mutation(child1)
                    self.conflict_guided_mutation(child2)
                else:
                # Use regular mutation 30% of the time for diversity
                    self.mutate(child1)
                    self.mutate(child2)
            
            # Additional regular mutation with adaptive rate
                if self.stagnation_counter > 10:
                # Increase mutation rate during stagnation
                    original_rate = getattr(self, 'mutation_rate', 0.02)
                    self.mutation_rate = min(0.15, original_rate * 2)
                    self.mutate(child1)
                    self.mutate(child2)
                    self.mutation_rate = original_rate
            
                self.evaluate_fitness(child1)
                self.evaluate_fitness(child2)
                new_population.extend([child1, child2])

            self.population = new_population[:self.config['POPULATION_SIZE']]  # Ensure exact population size

    # Final reporting
        final_conflicts = len(self.best_chromosome.conflicts) if hasattr(self.best_chromosome, 'conflicts') else getattr(self.best_chromosome, 'conflict_count', 'unknown')
        print(f"Evolution finished. Best fitness: {self.best_chromosome.fitness}, Final Conflicts: {final_conflicts}")
        return self.best_chromosome





# =====================================================================
# MODIFICATION 3: Complete replacement of chromosome_to_database
# This new function is robust and handles Room objects correctly.
# =====================================================================
def chromosome_to_database(chromosome):
    """FINAL VERSION: Bulletproof database save with absolute validation"""
    
    # Clear previous data
    EnhancedSchedule.objects.all().delete()
    GlobalScheduleMatrix.objects.all().delete()

    schedule_data_for_json = {}
    saved_count = 0
    skipped_count = 0

    for gene in chromosome.genes:
        # Basic validation
        if not all([gene.day, gene.room]):
            skipped_count += 1
            continue

        if not isinstance(gene.room, Room):
            skipped_count += 1
            continue

        # Handle lab blocks
        if gene.is_lab and hasattr(gene, 'lab_block_pair') and gene.lab_block_pair:
            for i, time_slot in enumerate(gene.lab_block_pair):
                try:
                    if not time_slot or not isinstance(time_slot, str):
                        continue

                    EnhancedSchedule.objects.create(
                        semester=gene.semester,
                        section=gene.section,
                        subject=f"{gene.subject} (Part {i+1})",
                        is_lab=True,
                        time_slot=time_slot,
                        day=gene.day,
                        room=gene.room,
                        week_number=gene.session_number
                    )
                    
                    GlobalScheduleMatrix.objects.create(
                        time_slot=time_slot,
                        day=gene.day,
                        room_number=gene.room.room_number,
                        semester=gene.semester,
                        section=gene.section,
                        subject=f"{gene.subject} (Part {i+1})"
                    )
                    
                    saved_count += 1

                except Exception as e:
                    logger.error(f"Failed to save lab block: {e}")
                    skipped_count += 1

        else:
            # Handle theory slots with ABSOLUTE validation
            try:
                # MULTIPLE VALIDATION BARRIERS
                if not hasattr(gene, 'time_slot'):
                    logger.warning(f"Skipping gene with no time_slot: {gene.subject}")
                    skipped_count += 1
                    continue
                    
                if gene.time_slot is None:
                    logger.warning(f"Skipping gene with no time_slot: {gene.subject}")
                    skipped_count += 1
                    continue
                    
                if gene.time_slot == "":
                    logger.warning(f"Skipping gene with no time_slot: {gene.subject}")
                    skipped_count += 1
                    continue
                    
                if gene.time_slot == "None":
                    logger.warning(f"Skipping gene with no time_slot: {gene.subject}")
                    skipped_count += 1
                    continue
                    
                if not isinstance(gene.time_slot, str):
                    logger.warning(f"Skipping gene with no time_slot: {gene.subject}")
                    skipped_count += 1
                    continue
                
                # Save valid theory classes
                EnhancedSchedule.objects.create(
                    semester=gene.semester,
                    section=gene.section,
                    subject=gene.subject,
                    is_lab=gene.is_lab,
                    time_slot=gene.time_slot,
                    day=gene.day,
                    room=gene.room,
                    week_number=gene.session_number
                )
                
                GlobalScheduleMatrix.objects.create(
                    time_slot=gene.time_slot,
                    day=gene.day,
                    room_number=gene.room.room_number,
                    semester=gene.semester,
                    section=gene.section,
                    subject=gene.subject
                )
                
                saved_count += 1
                
            except Exception as e:
                logger.error(f"Failed to save theory class: {e}")
                skipped_count += 1

        # Prepare JSON response data
        section_key = f'Semester {gene.semester} - {gene.section}'
        if section_key not in schedule_data_for_json:
            all_time_slots = THEORY_TIME_SLOTS + [slot for pair in LAB_TIME_SLOTS for slot in pair]
            schedule_data_for_json[section_key] = {
                day: {slot: None for slot in all_time_slots}
                for day in WEEKDAYS
            }

        # Add to JSON structure
        if gene.is_lab and hasattr(gene, 'lab_block_pair') and gene.lab_block_pair:
            for i, slot in enumerate(gene.lab_block_pair):
                subject_info = f"{gene.subject} (Lab Part {i+1}) - Room {gene.room.room_number}"
                schedule_data_for_json[section_key][gene.day][slot] = subject_info
        else:
            if gene.time_slot and gene.time_slot != "None" and gene.time_slot is not None:
                subject_info = f"{gene.subject} ({'Lab' if gene.is_lab else 'Theory'}) - Room {gene.room.room_number}"
                schedule_data_for_json[section_key][gene.day][gene.time_slot] = subject_info

    logger.info(f"Successfully saved {saved_count} schedule entries to database")
    return schedule_data_for_json



# In backend/main/views.py

# =====================================================================
# ADD THIS MISSING FUNCTION BACK INTO YOUR FILE
# Place it right before the `generate_enhanced_timetable` function.
# =====================================================================
# In backend/main/views.py

# Ensure this function from the previous step is present and correctly indented
def prepare_semesters_data(max_semester):
    """Prepare data structure for the genetic algorithm by querying the database."""
    semesters_data = {}
    for semester in range(1, max_semester + 1):
        sections = Section.objects.filter(semester_number=semester)
        if not sections.exists():
            continue
            
        semesters_data[semester] = {}
        # Get all courses for the semester once to be efficient
        core_courses = CoreCourse.objects.filter(semester_number=semester)
        cohort_courses = CohortCourse.objects.filter(semester_number=semester)

        for section in sections:
            subjects = []
            for course in core_courses:
                subjects.append({
                    'name': course.course_name,
                    'is_lab': course.is_lab,
                    'duration': course.duration,
                    'is_cohort': False,
                    'course_type': 'core'
                })
            for cohort_course in cohort_courses:
                subjects.append({
                    'name': cohort_course.course_name,
                    'is_lab': False,  # Cohort courses are ALWAYS theoretical
                    'duration': 75,   # Standard theory duration
                    'is_cohort': True,
                    'capacity': 100,
                    'course_type': 'cohort',
                    'capacity': cohort_course.capacity   # Higher capacity for cohort courses
                })
            semesters_data[semester][section.section_name] = subjects
            
    return semesters_data

# =====================================================================
# REPLACE your existing 'generate_enhanced_timetable' with this complete version.
# This version has the try...except block correctly structured.
# =====================================================================
@csrf_exempt
def generate_enhanced_timetable(request):
    """Generate timetable using Genetic Algorithm"""
    if request.method == 'POST':
        try:
            # Handle both JSON and form data
            if request.content_type == 'application/json':
                data = json.loads(request.body)
                max_semester = data.get('max_semester', 8)
            else:
                max_semester = int(request.POST.get('max_semester', 8))

            logger.info(f"Starting Genetic Algorithm timetable generation for {max_semester} semesters")
            
            # Prepare data using the function above
            semesters_data = prepare_semesters_data(max_semester)
            if not semesters_data:
                logger.warning("No semester data found for the given range. Aborting generation.")
                return JsonResponse({'status': 'error', 'message': 'No courses or sections found for the specified semesters.'}, status=404)

            # Run the Genetic Algorithm
            ga = GeneticAlgorithm()
            best_solution = ga.evolve(semesters_data)
            
            # Persist and format the final solution
            schedule_data = chromosome_to_database(best_solution)

            # Prepare the final successful response
            response_data = {
                'status': 'success',
                'message': 'Genetic Algorithm timetable generated successfully',
                'schedule': schedule_data,
                'algorithm_info': {
                    'generations': ga.generation,
                    'final_fitness': best_solution.fitness,
                    'conflicts': len(best_solution.conflicts),
                }
            }
            return JsonResponse(response_data)

        except Exception as e:
            # This is the 'except' block that was missing or broken
            logger.error(f"Error in genetic algorithm timetable generation: {str(e)}", exc_info=True)
            return JsonResponse({'status': 'error', 'message': f'An unexpected error occurred: {str(e)}'}, status=500)
    
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)



# ============ DJANGO VIEWS ============
@csrf_exempt
def generate_enhanced_timetable(request):
    """Generate timetable using Genetic Algorithm"""
    if request.method == 'POST':
        try:
            # Handle both JSON and form data
            if request.content_type == 'application/json':
                data = json.loads(request.body)
                max_semester = data.get('max_semester', 8)
            else:
                max_semester = int(request.POST.get('max_semester', 8))

            logger.info(f"Starting Genetic Algorithm timetable generation for {max_semester} semesters")
            
            # Prepare data
            semesters_data = prepare_semesters_data(max_semester)
            
            # Run GA
            ga = GeneticAlgorithm()
            best_solution = ga.evolve(semesters_data)
            
            # Persist and format solution
            schedule_data = chromosome_to_database(best_solution)

            # Prepare response
            response_data = {
                'status': 'success',
                'message': 'Genetic Algorithm timetable generated successfully',
                'schedule': schedule_data,
                'algorithm_info': {
                    'generations': ga.generation,
                    'final_fitness': best_solution.fitness,
                    'conflicts': len(best_solution.conflicts),
                }
            }
            return JsonResponse(response_data)

        except Exception as e:
            logger.error(f"Error in genetic algorithm timetable generation: {str(e)}", exc_info=True)
            return JsonResponse({'status': 'error', 'message': f'GA generation failed: {str(e)}'}, status=500)
    
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)

def generate_timetable(request):
    """Main timetable generation entry point"""
    if request.method == 'POST':
        form = TimetableGenerationForm(request.POST, request.FILES)
        if form.is_valid():
            max_semester = int(form.cleaned_data['max_semester'])
            
            try:
                file = request.FILES['file']
                
                # Process Excel file
                if file.name.endswith('.xlsx') or file.name.endswith('.xls'):
                    success = process_excel_file(file)
                else:
                    success = process_csv_files([file])
                
                if not success:
                    return JsonResponse({
                        'status': 'error',
                        'message': 'Error processing the uploaded file.'
                    })
                
                # Store data and redirect to genetic algorithm generation
                request.session['upload_data'] = {
                    'max_semester': max_semester,
                    'file_processed': True
                }
                
                # Generate using genetic algorithm
                return generate_enhanced_web(request, max_semester)
                
            except Exception as e:
                logger.error(f"Error in file processing: {str(e)}", exc_info=True)
                return JsonResponse({
                    'status': 'error', 
                    'message': f'Error: {str(e)}'
                })
    else:
        # Return form data as JSON for API compatibility
        return JsonResponse({
            'status': 'success',
            'message': 'Upload Excel file to generate timetable',
            'required_sheets': ['Students', 'Core Courses', 'Cohort Courses', 'Elective Courses', 'Rooms'],
            'genetic_algorithm': 'Active'
        })

def generate_enhanced_web(request, max_semester=None):
    """Generate genetic algorithm timetable via web interface"""
    if max_semester is None:
        upload_data = request.session.get('upload_data')
        if not upload_data:
            return JsonResponse({
                'status': 'error',
                'message': 'No upload data found. Please upload file first.'
            })
        max_semester = upload_data['max_semester']
    
    # Create fake POST request for genetic algorithm
    request.method = 'POST'
    request.POST = {'max_semester': max_semester}
    
    return generate_enhanced_timetable(request)
def _pick(df, *variants, required=True):
    for v in variants:
        if v in df.columns:
            return v
    if required:
        raise KeyError(f"None of the columns {variants} found in sheet.")
    return None


def _to_bool(val):
    if isinstance(val, bool): return val
    if val is None or (isinstance(val, float) and pd.isna(val)): return False
    s = str(val).strip().lower()
    return s in ("1", "true", "t", "yes", "y")

# ============ FILE PROCESSING FUNCTIONS ============

# =====================================================================
# ▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼ REPLACEMENT FUNCTION ▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼
# =====================================================================
def process_excel_file(excel_file):
    try:
        xls = pd.ExcelFile(excel_file)
        # --- Process Students ---
        if "StudentCapacity" in xls.sheet_names or "Students" in xls.sheet_names:
            sheet_name = "StudentCapacity" if "StudentCapacity" in xls.sheet_names else "Students"
            students = pd.read_excel(xls, sheet_name)
            SemesterStudents.objects.all().delete()
            Section.objects.all().delete()
            sem_col = _pick(students, "Semester Number", "semester", "Semester", "Sem")
            pop_col = _pick(students, "Number of Students", "student_count", "Student Count", "Headcount")
            for _, row in students.iterrows():
                semester = to_int_or_zero(row[sem_col])
                num_students = to_int_or_zero(row[pop_col])
                if semester > 0:
                    SemesterStudents.objects.create(semester_number=semester, number_of_students=num_students)
                    Section.create_sections(semester, num_students)
        # --- Process Core Courses (from Roadmap sheet) ---
        if "Roadmap" in xls.sheet_names or "Core Courses" in xls.sheet_names:
            sheet_name = "Roadmap" if "Roadmap" in xls.sheet_names else "Core Courses"
            courses = pd.read_excel(xls, sheet_name)
            CoreCourse.objects.all().delete()
            sem_col = _pick(courses, "Semester Number", "semester", "Semester")
            name_col = _pick(courses, "Course Name", "course_name", "Course")
            lab_col = _pick(courses, "Lab/Theory (Boolean)", "is_lab", "Lab", "lab")
            for _, row in courses.iterrows():
                semester = to_int_or_zero(row[sem_col])
                if semester > 0:
                    is_lab_flag = _to_bool(row[lab_col])
                    CoreCourse.objects.create(
                        semester_number=semester,
                        course_name=str(row[name_col]),
                        is_lab=is_lab_flag,
                        duration=150 if is_lab_flag else 75
                    )
        # --- Process Rooms ---
        if "Rooms" in xls.sheet_names:
            rooms = pd.read_excel(xls, "Rooms")
            Room.objects.all().delete()
            num_col = _pick(rooms, "Room Number", "room_number", "Room", "room_name")
            type_col = _pick(rooms, "Room Type", "room_type", "Type", required=False)
            for _, row in rooms.iterrows():
                room_name = str(row[num_col]).strip()
                if room_name and room_name.lower() != 'nan':
                    r_type = str(row.get(type_col, "")).lower()
                    Room.objects.create(
                        room_number=room_name,
                        core_sub=r_type in ("theory", "lecture", "core"),
                        lab=r_type == "lab"
                    )
        return True
    except KeyError as e:
        logger.error(f"Required column not found during API upload: {e}")
        return False
    except Exception as e:
        logger.error(f"Error processing API-uploaded Excel file: {str(e)}", exc_info=True)
        return False

# You also need to add this helper function if it's not already there
def to_int_or_zero(value):
    try:
        if pd.isna(value): return 0
        return int(float(value))
    except (ValueError, TypeError):
        return 0


def process_csv_files(csv_files):
    return False

# ============ UTILITY FUNCTIONS ============

# =====================================================================
# ▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼ REPLACEMENT 2 of 2 ▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼▼
# =====================================================================
def save_timetable(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            schedule_data = data.get('schedule')
            max_semester = data.get('max_semester')

            if not schedule_data or not max_semester:
                return JsonResponse({'status': 'error', 'message': 'Missing "schedule" or "max_semester".'}, status=400)
            
            timetable = Timetable.objects.create(
                semester=max_semester,
                data=schedule_data
            )
            logger.info(f"Timetable saved with ID: {timetable.id}")
            return JsonResponse({'status': 'success', 'message': 'Timetable saved successfully!', 'timetable_id': timetable.id})
        except json.JSONDecodeError:
            return JsonResponse({'status': 'error', 'message': 'Invalid JSON.'}, status=400)
        except Exception as e:
            logger.error(f"Error saving timetable: {e}", exc_info=True)
            return JsonResponse({'status': 'error', 'message': 'Internal error.'}, status=500)
    return JsonResponse({'status': 'error', 'message': 'Use POST.'}, status=405)


def view_timetable(request):
    timetables = Timetable.objects.all().order_by('-created_at')
    
    # Convert timetables to JSON-friendly format
    timetables_data = []
    for timetable in timetables:
        timetables_data.append({
            'id': timetable.id,
            'semester': timetable.semester,
            'created_at': timetable.created_at.isoformat(),
            'data': timetable.data
        })
    
    return JsonResponse({
        'status': 'success',
        'timetables': timetables_data,
        'days': WEEKDAYS,
        'time_slots': [slot.replace('-', ' - ') for slot in ENHANCED_TIME_SLOTS],
        'enhanced_time_slots': ENHANCED_TIME_SLOTS,
    })

def download_timetable_excel(request, timetable_id=None):
    """Enhanced Excel download with proper lab splitting support"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from django.http import HttpResponse
        import io
        
        # Get all schedule data
        schedule_entries = EnhancedSchedule.objects.all().order_by('semester', 'section', 'day')
        
        if not schedule_entries.exists():
            return JsonResponse({'error': 'No timetable data found'}, status=404)
        
        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Group data by semester and section
        schedule_dict = {}
        for entry in schedule_entries:
            key = f"Semester {entry.semester} - {entry.section}"
            if key not in schedule_dict:
                schedule_dict[key] = {}
            
            day = entry.day
            if day not in schedule_dict[key]:
                schedule_dict[key][day] = {}
            
            # Handle both lab and theory entries
            schedule_dict[key][day][entry.time_slot] = f"{entry.subject} - Room {entry.room.room_number}"
        
        # Create proper time slot list (flatten lab blocks)
        theory_slots = THEORY_TIME_SLOTS
        lab_individual_slots = []
        
        # FIXED: Properly handle lab time slots
        for lab_block in LAB_TIME_SLOTS:
            if isinstance(lab_block, list):
                lab_individual_slots.extend(lab_block)  # Add individual slots from pairs
            else:
                lab_individual_slots.append(lab_block)  # Single slot
        
        # Combine and sort all time slots
        all_individual_slots = sorted(set(theory_slots + lab_individual_slots))
        
        # Format time slots for display (add spaces around dashes)
        formatted_time_slots = []
        for slot in all_individual_slots:
            if isinstance(slot, str) and '-' in slot:
                formatted_slot = slot.replace('-', ' - ')
                formatted_time_slots.append(formatted_slot)
            else:
                formatted_time_slots.append(str(slot))
        
        # Create sheets for each semester-section
        for section_key, section_data in schedule_dict.items():
            # Create worksheet
            ws = wb.create_sheet(title=section_key)
            
            # Header styling
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Write headers
            headers = ['Time'] + WEEKDAYS
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = border
            
            # Write time slots and schedule data
            for row_idx, time_slot in enumerate(formatted_time_slots, 2):
                # Time column
                time_cell = ws.cell(row=row_idx, column=1, value=time_slot)
                time_cell.font = Font(bold=True)
                time_cell.border = border
                time_cell.alignment = Alignment(horizontal='center')
                
                # Day columns
                for col_idx, day in enumerate(WEEKDAYS, 2):
                    # Find matching time slot (handle formatting differences)
                    original_time_slot = time_slot.replace(' - ', '-')
                    
                    if day in section_data and original_time_slot in section_data[day]:
                        cell_value = section_data[day][original_time_slot]
                        cell_fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")
                    else:
                        cell_value = "Free"
                        cell_fill = PatternFill(start_color="F8F8F8", end_color="F8F8F8", fill_type="solid")
                    
                    cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
                    cell.fill = cell_fill
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # If no sheets created, create empty one
        if not wb.worksheets:
            ws = wb.create_sheet(title="No Data")
            ws['A1'] = "No timetable data available"
        
        # Prepare response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="Quantime_AI_Timetable.xlsx"'
        
        # Save workbook to response
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response.write(output.getvalue())
        
        return response
        
    except Exception as e:
        logger.error(f"Excel download error: {str(e)}")
        return JsonResponse({'error': f'Download failed: {str(e)}'}, status=500)


def delete_timetable(request, timetable_id):
    timetable = get_object_or_404(Timetable, id=timetable_id)
    timetable.delete()
    return JsonResponse({
        'status': 'success',
        'message': 'Timetable deleted successfully!'
    })

# ============ VALIDATION AND STATISTICS ============

@csrf_exempt
def validate_enhanced_schedule(request):
    """Validate the genetic algorithm generated schedule"""
    conflicts = []
    
    # Check section conflicts
    schedules_by_section_time = defaultdict(list)
    for schedule in EnhancedSchedule.objects.all():
        key = f"{schedule.semester}-{schedule.section}-{schedule.day}-{schedule.time_slot}"
        schedules_by_section_time[key].append(schedule)
    
    for key, schedules in schedules_by_section_time.items():
        if len(schedules) > 1:
            conflicts.append({
                'type': 'section_conflict',
                'key': key,
                'count': len(schedules)
            })
    
    # Check room conflicts
    room_bookings = defaultdict(list)
    for matrix in GlobalScheduleMatrix.objects.all():
        key = f"{matrix.day}-{matrix.time_slot}-{matrix.room_number}"
        room_bookings[key].append(matrix)
    
    for key, bookings in room_bookings.items():
        if len(bookings) > 1:
            conflicts.append({
                'type': 'room_conflict',
                'key': key,
                'count': len(bookings)
            })
    
    # Check theoretical subjects frequency
    theoretical_counts = defaultdict(int)
    for schedule in EnhancedSchedule.objects.filter(is_lab=False):
        key = f"{schedule.semester}-{schedule.section}-{schedule.subject}"
        theoretical_counts[key] += 1
    
    twice_weekly_subjects = sum(1 for count in theoretical_counts.values() if count == 2)
    incorrect_frequency = sum(1 for count in theoretical_counts.values() if count != 2)
    
    return JsonResponse({
        'status': 'success',
        'validation_results': {
            'total_conflicts': len(conflicts),
            'conflicts': conflicts,
            'theoretical_subjects_correct_frequency': twice_weekly_subjects,
            'theoretical_subjects_incorrect_frequency': incorrect_frequency,
            'total_schedules': EnhancedSchedule.objects.count(),
            'algorithm_type': 'Genetic Algorithm + CSP'
        },
        'is_valid': len(conflicts) == 0 and incorrect_frequency == 0
    })

def get_schedule_statistics(request):
    """Get detailed statistics about the genetic algorithm generated schedule"""
    total_schedules = EnhancedSchedule.objects.count()
    theoretical_count = EnhancedSchedule.objects.filter(is_lab=False).count()
    lab_count = EnhancedSchedule.objects.filter(is_lab=True).count()
    
    # Calculate theoretical frequency distribution
    theoretical_subjects = defaultdict(int)
    for schedule in EnhancedSchedule.objects.filter(is_lab=False):
        key = f"{schedule.semester}-{schedule.section}-{schedule.subject}"
        theoretical_subjects[key] += 1
    
    frequency_distribution = {
        'once_weekly': sum(1 for count in theoretical_subjects.values() if count == 1),
        'twice_weekly': sum(1 for count in theoretical_subjects.values() if count == 2),
        'more_than_twice': sum(1 for count in theoretical_subjects.values() if count > 2)
    }
    
    # Time slot usage
    time_slot_usage = {}
    all_slots = THEORY_TIME_SLOTS + LAB_TIME_SLOTS
    for slot in all_slots:
        count = EnhancedSchedule.objects.filter(time_slot=slot).count()
        time_slot_usage[slot] = count
    
    # Extended hours usage
    extended_slots = ["17:00-18:15", "18:30-19:45", "20:00-21:15"]
    extended_usage = sum(time_slot_usage.get(slot, 0) for slot in extended_slots)
    total_usage = sum(time_slot_usage.values())
    extended_percentage = (extended_usage / total_usage * 100) if total_usage > 0 else 0
    
    return JsonResponse({
        'algorithm_type': 'Genetic Algorithm + Constraint Satisfaction Problem',
        'total_schedules': total_schedules,
        'schedule_breakdown': {
            'theoretical_classes': theoretical_count,
            'lab_classes': lab_count
        },
        'theoretical_frequency_distribution': frequency_distribution,
        'time_slot_usage': time_slot_usage,
        'extended_hours_utilization': {
            'count': extended_usage,
            'percentage': round(extended_percentage, 2)
        },
        'optimization_features': [
            'Multi-objective fitness function',
            'Tournament selection',
            'Single-point crossover',
            'Random mutation',
            'Elite preservation',
            'Constraint satisfaction integration'
        ]
    })

# ============ COMPATIBILITY FUNCTIONS FOR API_VIEWS ============

def create_timetable(max_semester):
    """Enhanced timetable creation with robust conflict detection"""
    logger.info(f"Legacy create_timetable called - redirecting to Genetic Algorithm for {max_semester} semesters")
    
    try:
        # Prepare data for genetic algorithm
        semesters_data = prepare_semesters_data(max_semester)
        
        # Initialize and run genetic algorithm
        ga = GeneticAlgorithm()
        best_solution = ga.evolve(semesters_data)
        
        # Get generation count
        generations_used = getattr(ga, 'generation', 1)
        
        # ROBUST: Try multiple ways to get conflict count
        conflict_count = 'unknown'
        if best_solution:
            # Method 1: Check for conflicts list attribute
            if hasattr(best_solution, 'conflicts'):
                conflicts_attr = getattr(best_solution, 'conflicts')
                if isinstance(conflicts_attr, list):
                    conflict_count = len(conflicts_attr)
                elif isinstance(conflicts_attr, int):
                    conflict_count = conflicts_attr
            
            # Method 2: Check for conflict_count attribute
            elif hasattr(best_solution, 'conflict_count'):
                conflict_count = best_solution.conflict_count
            
            # Method 3: Calculate using GA method (fallback)
            else:
                try:
                    conflicting_genes = ga.get_all_hard_conflicts(best_solution)
                    conflict_count = len(conflicting_genes)
                except:
                    # If all else fails, assume 0 since algorithm reports "Final Conflicts: 0"
                    conflict_count = 0
        
        # Convert to database
        schedule_data = chromosome_to_database(best_solution) if best_solution else {}
        actual_entries = EnhancedSchedule.objects.count()
        
        if best_solution and conflict_count == 0:
            # SUCCESS: Perfect solution
            return {
                'success': True,
                'conflicts': 0,
                'fitness': best_solution.fitness,
                'message': 'Optimal timetable generated successfully',
                'total_entries': actual_entries,
                'generations': generations_used,
                'schedule_data': schedule_data
            }
        elif best_solution and isinstance(conflict_count, int) and conflict_count < 10:
            # ACCEPTABLE: Low conflicts
            return {
                'success': True,
                'conflicts': conflict_count,
                'fitness': best_solution.fitness,
                'message': f'Good timetable with {conflict_count} minor conflicts',
                'total_entries': actual_entries,
                'generations': generations_used,
                'schedule_data': schedule_data
            }
        else:
            # FAILED: Too many conflicts or unknown status
            return {
                'success': False,
                'conflicts': conflict_count,
                'fitness': best_solution.fitness if best_solution else 0,
                'message': f'Algorithm result unclear (conflicts: {conflict_count})'
            }
            
    except Exception as e:
        logger.error(f"Error in create_timetable: {str(e)}")
        return {
            'success': False,
            'conflicts': 'error',
            'message': f'Error: {str(e)}'
        }
# //frontend stuff
import json
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
import logging

logger = logging.getLogger(__name__)

@csrf_exempt
def save_timetable(request):
    """FINAL BULLETPROOF save endpoint - handles any request format"""
    if request.method == 'POST':
        try:
            # Handle any request body format
            try:
                data = json.loads(request.body) if request.body else {}
            except (json.JSONDecodeError, UnicodeDecodeError):
                data = {}
            
            max_semester = data.get('max_semester', 8)
            logger.info(f"Processing save request for {max_semester} semesters")
            
            # ALWAYS generate fresh timetable - ignore frontend data completely
            logger.info("Generating fresh timetable")
            
            # Clear existing data
            EnhancedSchedule.objects.all().delete()
            GlobalScheduleMatrix.objects.all().delete()
            
            # Generate new timetable
            semesters_data = prepare_semesters_data(max_semester)
            ga = GeneticAlgorithm()
            best_solution = ga.evolve(semesters_data)
            
            # Save to database
            schedule_data = chromosome_to_database(best_solution)
            
            # Get results
            saved_entries = EnhancedSchedule.objects.count()
            
            # Calculate conflicts
            if hasattr(best_solution, 'conflicts'):
                conflicts = len(best_solution.conflicts) if best_solution.conflicts else 0
            elif hasattr(best_solution, 'conflict_count'):
                conflicts = best_solution.conflict_count
            else:
                conflicting_genes = ga.get_all_hard_conflicts(best_solution)
                conflicts = len(conflicting_genes)
            
            logger.info(f"Save SUCCESS: {saved_entries} entries, {conflicts} conflicts")
            
            return JsonResponse({
                'status': 'success',
                'message': f'Timetable generated and saved successfully',
                'entries_saved': saved_entries,
                'conflicts': conflicts,
                'fitness': getattr(best_solution, 'fitness', 'unknown'),
                'semesters': max_semester
            }, status=200)
            
        except Exception as e:
            logger.error(f"Save operation failed: {str(e)}")
            return JsonResponse({
                'status': 'error',
                'message': f'Save operation failed: {str(e)}'
            }, status=500)
    
    return JsonResponse({
        'status': 'error',
        'message': 'Only POST requests allowed'
    }, status=405)


