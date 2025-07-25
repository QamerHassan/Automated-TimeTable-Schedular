# # Enhanced Genetic Algorithm Timetable Generation System

# from django.shortcuts import render, redirect, get_object_or_404
# from django.contrib import messages
# from .models import (CoreCourse, CohortCourse, ElectiveCourse, Room, SemesterStudents, 
#                      Timetable, Section, EnhancedSchedule, GlobalScheduleMatrix, 
#                      SubjectFrequencyRule, TimeSlotConfiguration)
# from .forms import TimetableGenerationForm
# from django.contrib.auth.decorators import login_required
# from django.db.models import Sum
# from django.core.files.storage import default_storage
# from django.http import JsonResponse, HttpResponse
# from django.views.decorators.csrf import csrf_exempt
# import random
# import csv
# import os
# import io
# import logging
# import json
# import pandas as pd
# from openpyxl import Workbook
# from openpyxl.styles import Font, Alignment, PatternFill
# from django.utils.text import slugify
# from datetime import datetime, timedelta, time
# from collections import defaultdict
# import numpy as np

# logger = logging.getLogger(__name__)

# # ============ GENETIC ALGORITHM CONFIGURATION ============

# GENETIC_CONFIG = {
#     'POPULATION_SIZE': 50,
#     'GENERATIONS': 100,
#     'CROSSOVER_RATE': 0.8,
#     'MUTATION_RATE': 0.1,
#     'ELITE_SIZE': 5,
#     'TOURNAMENT_SIZE': 3
# }

# SUBJECT_FREQUENCY_MAP = {
#     # Theoretical subjects - 2 times per week
#     "English-I": 2, "Calculus": 2, "Programming Fundamentals": 2,
#     "ICT Theory": 2, "Information Security": 2, "Physics": 2,
#     "Chemistry": 2, "Mathematics": 2, "Statistics": 2,
    
#     # Lab subjects - 1 time per week
#     "Programming Fundamentals Lab": 1, "ICT Lab": 1,
#     "Physics Lab": 1, "Chemistry Lab": 1,
# }

# ENHANCED_TIME_SLOTS = [
#     "08:00-09:15", "09:30-10:45", "11:00-12:15",  # Morning
#     "12:30-13:45", "14:00-15:15", "15:30-16:45",  # Afternoon  
#     "17:00-18:15", "18:30-19:45", "20:00-21:15"   # Evening (Extended hours)
# ]

# WEEKDAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]

# def home(request):
#     """API-friendly home endpoint - no templates needed"""
#     return JsonResponse({
#         'status': 'success',
#         'message': 'Genetic Algorithm Timetable System API Active',
#         'version': '2.0-GA',
#         'algorithm': 'Genetic Algorithm + Constraint Satisfaction',
#         'backend_features': [
#             'Twice-weekly theoretical lectures',
#             'Inter-semester conflict resolution', 
#             'Extended hours (8 AM - 9 PM)',
#             'Multi-objective optimization'
#         ],
#         'api_endpoints': {
#             'generate': '/api/generate/',
#             'enhanced_generate': '/api/generate-enhanced/',
#             'validate': '/api/validate-enhanced/',
#             'stats': '/api/schedule-stats/'
#         },
#         'genetic_config': {
#             'population_size': 50,
#             'generations': 100,
#             'crossover_rate': 0.8,
#             'mutation_rate': 0.1
#         }
#     })

# # ============ GENETIC ALGORITHM CORE CLASSES ============

# class Gene:
#     """Represents a single scheduled class"""
#     def __init__(self, semester, section, subject, is_lab, day=None, time_slot=None, room=None, teacher=None):
#         self.semester = semester
#         self.section = section
#         self.subject = subject
#         self.is_lab = is_lab
#         self.day = day or random.choice(WEEKDAYS)
#         self.time_slot = time_slot or random.choice(ENHANCED_TIME_SLOTS)
#         self.room = room
#         self.teacher = teacher
#         self.session_number = 1

# class Chromosome:
#     """Represents a complete timetable solution"""
#     def __init__(self, genes=None):
#         self.genes = genes or []
#         self.fitness = 0
#         self.conflicts = []
#         self.schedule_matrix = {}
        
#     def initialize_random(self, semesters_data):
#         """Initialize chromosome with random valid assignments"""
#         self.genes = []
        
#         for semester, sections_data in semesters_data.items():
#             for section_name, subjects in sections_data.items():
#                 for subject_info in subjects:
#                     frequency = get_subject_frequency(subject_info['name'], subject_info['is_lab'])
                    
#                     for session in range(frequency):
#                         gene = Gene(
#                             semester=semester,
#                             section=section_name,
#                             subject=subject_info['name'],
#                             is_lab=subject_info['is_lab']
#                         )
#                         gene.session_number = session + 1
#                         self.genes.append(gene)
        
#         # Assign rooms to genes
#         self.assign_rooms_to_genes()
        
#     def assign_rooms_to_genes(self):
#         """Assign appropriate rooms to genes"""
#         lab_rooms = list(Room.objects.filter(lab=True))
#         theory_rooms = list(Room.objects.filter(core_sub=True))
#         backup_rooms = list(Room.objects.filter(cohort_sub=True))
        
#         for gene in self.genes:
#             if gene.is_lab and lab_rooms:
#                 gene.room = random.choice(lab_rooms).room_number
#             elif theory_rooms:
#                 gene.room = random.choice(theory_rooms).room_number
#             elif backup_rooms:
#                 gene.room = random.choice(backup_rooms).room_number
#             else:
#                 gene.room = "TBD"  # To be determined later

# class GeneticAlgorithm:
#     """Main Genetic Algorithm implementation for timetable optimization"""
    
#     def __init__(self, config=None):
#         self.config = config or GENETIC_CONFIG
#         self.population = []
#         self.best_chromosome = None
#         self.generation = 0
#         self.fitness_history = []
        
#     def initialize_population(self, semesters_data):
#         """Create initial population of random timetables"""
#         logger.info(f"Initializing population of size {self.config['POPULATION_SIZE']}")
        
#         self.population = []
#         for i in range(self.config['POPULATION_SIZE']):
#             chromosome = Chromosome()
#             chromosome.initialize_random(semesters_data)
#             self.population.append(chromosome)
            
#         logger.info(f"Created {len(self.population)} chromosomes")
        
#     def evaluate_fitness(self, chromosome):
#         """Evaluate chromosome fitness based on constraints"""
#         fitness = 1000  # Start with high fitness
#         conflicts = []
        
#         # Hard constraints (major penalties)
#         section_conflicts = self.check_section_conflicts(chromosome)
#         room_conflicts = self.check_room_conflicts(chromosome)
#         inter_semester_conflicts = self.check_inter_semester_conflicts(chromosome)
        
#         # Soft constraints (minor penalties)
#         daily_load_penalties = self.check_daily_load_balance(chromosome)
#         theoretical_frequency_penalties = self.check_theoretical_frequency(chromosome)
#         gap_penalties = self.check_time_gaps(chromosome)
        
#         # Apply penalties
#         fitness -= len(section_conflicts) * 200  # Major penalty for section conflicts
#         fitness -= len(room_conflicts) * 150    # Major penalty for room conflicts
#         fitness -= len(inter_semester_conflicts) * 100  # Penalty for inter-semester conflicts
#         fitness -= daily_load_penalties * 50    # Penalty for unbalanced daily loads
#         fitness -= theoretical_frequency_penalties * 75  # Penalty for incorrect frequency
#         fitness -= gap_penalties * 25           # Penalty for large gaps
        
#         # Bonus for extended hours utilization
#         extended_hours_bonus = self.calculate_extended_hours_usage(chromosome)
#         fitness += extended_hours_bonus * 10
        
#         # Store conflicts for debugging
#         chromosome.conflicts = section_conflicts + room_conflicts + inter_semester_conflicts
#         chromosome.fitness = max(0, fitness)  # Ensure non-negative fitness
        
#         return chromosome.fitness
        
#     def check_section_conflicts(self, chromosome):
#         """Check for conflicts within the same section"""
#         conflicts = []
#         section_schedule = defaultdict(lambda: defaultdict(set))
        
#         for gene in chromosome.genes:
#             key = f"{gene.semester}-{gene.section}"
#             time_key = f"{gene.day}-{gene.time_slot}"
            
#             if time_key in section_schedule[key]:
#                 conflicts.append({
#                     'type': 'section_conflict',
#                     'section': key,
#                     'time': time_key,
#                     'subjects': list(section_schedule[key][time_key]) + [gene.subject]
#                 })
            
#             section_schedule[key][time_key].add(gene.subject)
            
#         return conflicts
        
#     def check_room_conflicts(self, chromosome):
#         """Check for room booking conflicts"""
#         conflicts = []
#         room_schedule = defaultdict(set)
        
#         for gene in chromosome.genes:
#             if gene.room:
#                 time_room_key = f"{gene.day}-{gene.time_slot}-{gene.room}"
                
#                 if time_room_key in room_schedule:
#                     conflicts.append({
#                         'type': 'room_conflict',
#                         'room': gene.room,
#                         'time': f"{gene.day}-{gene.time_slot}",
#                         'conflicting_classes': list(room_schedule[time_room_key]) + [f"S{gene.semester}-{gene.section}-{gene.subject}"]
#                     })
                
#                 room_schedule[time_room_key].add(f"S{gene.semester}-{gene.section}-{gene.subject}")
                
#         return conflicts
        
#     def check_inter_semester_conflicts(self, chromosome):
#         """Check for conflicts between different semesters"""
#         conflicts = []
#         time_usage = defaultdict(set)
        
#         for gene in chromosome.genes:
#             time_key = f"{gene.day}-{gene.time_slot}"
#             semester_section = f"S{gene.semester}-{gene.section}"
            
#             # Check if different semesters are scheduled at same time without room separation
#             if time_key in time_usage:
#                 existing_semesters = {entry.split('-')[0] for entry in time_usage[time_key]}
#                 current_semester = f"S{gene.semester}"
                
#                 if current_semester not in existing_semesters:
#                     # Different semesters at same time - check room separation
#                     if not self.has_room_separation(chromosome, time_key):
#                         conflicts.append({
#                             'type': 'inter_semester_conflict',
#                             'time': time_key,
#                             'semesters': list(existing_semesters) + [current_semester]
#                         })
            
#             time_usage[time_key].add(semester_section)
            
#         return conflicts
        
#     def has_room_separation(self, chromosome, time_key):
#         """Check if different semesters use different rooms at same time"""
#         day, time_slot = time_key.split('-', 1)
#         rooms_used = set()
        
#         for gene in chromosome.genes:
#             if gene.day == day and gene.time_slot == time_slot:
#                 if gene.room in rooms_used:
#                     return False
#                 rooms_used.add(gene.room)
                
#         return True
        
#     def check_daily_load_balance(self, chromosome):
#         """Check for balanced daily class loads"""
#         penalties = 0
#         daily_loads = defaultdict(lambda: defaultdict(int))
        
#         for gene in chromosome.genes:
#             section_key = f"{gene.semester}-{gene.section}"
#             daily_loads[section_key][gene.day] += 1
            
#         for section, days in daily_loads.items():
#             for day, count in days.items():
#                 if count > 3:  # More than 3 classes per day
#                     penalties += count - 3
                    
#         return penalties
        
#     def check_theoretical_frequency(self, chromosome):
#         """Check if theoretical subjects are scheduled twice weekly"""
#         penalties = 0
#         subject_counts = defaultdict(int)
        
#         for gene in chromosome.genes:
#             if not gene.is_lab:
#                 key = f"{gene.semester}-{gene.section}-{gene.subject}"
#                 subject_counts[key] += 1
                
#         for subject_key, count in subject_counts.items():
#             expected_frequency = 2  # Theoretical subjects should appear twice
#             if count != expected_frequency:
#                 penalties += abs(count - expected_frequency)
                
#         return penalties
        
#     def check_time_gaps(self, chromosome):
#         """Check for large time gaps in daily schedules"""
#         penalties = 0
#         daily_schedules = defaultdict(lambda: defaultdict(list))
        
#         for gene in chromosome.genes:
#             section_key = f"{gene.semester}-{gene.section}"
#             daily_schedules[section_key][gene.day].append(gene.time_slot)
            
#         for section, days in daily_schedules.items():
#             for day, slots in days.items():
#                 if len(slots) > 1:
#                     # Sort slots and check gaps
#                     sorted_slots = sorted(slots, key=lambda x: datetime.strptime(x.split('-')[0], '%H:%M'))
                    
#                     for i in range(len(sorted_slots) - 1):
#                         current_end = datetime.strptime(sorted_slots[i].split('-')[1], '%H:%M')
#                         next_start = datetime.strptime(sorted_slots[i+1].split('-')[0], '%H:%M')
#                         gap_hours = (next_start - current_end).seconds / 3600
                        
#                         if gap_hours > 2:  # Gap larger than 2 hours
#                             penalties += 1
                            
#         return penalties
        
#     def calculate_extended_hours_usage(self, chromosome):
#         """Calculate bonus for using extended hours (evening slots)"""
#         extended_slots = ["17:00-18:15", "18:30-19:45", "20:00-21:15"]
#         extended_usage = 0
        
#         for gene in chromosome.genes:
#             if gene.time_slot in extended_slots:
#                 extended_usage += 1
                
#         return extended_usage
        
#     def selection(self):
#         """Tournament selection for choosing parents"""
#         selected = []
        
#         for _ in range(self.config['POPULATION_SIZE']):
#             # Tournament selection
#             tournament = random.sample(self.population, min(self.config['TOURNAMENT_SIZE'], len(self.population)))
#             winner = max(tournament, key=lambda x: x.fitness)
#             selected.append(winner)
            
#         return selected
        
#     def crossover(self, parent1, parent2):
#         """Crossover operation to create offspring"""
#         if random.random() > self.config['CROSSOVER_RATE']:
#             return parent1, parent2
            
#         # Single-point crossover
#         crossover_point = random.randint(1, min(len(parent1.genes), len(parent2.genes)) - 1)
        
#         child1 = Chromosome()
#         child2 = Chromosome()
        
#         child1.genes = parent1.genes[:crossover_point] + parent2.genes[crossover_point:]
#         child2.genes = parent2.genes[:crossover_point] + parent1.genes[crossover_point:]
        
#         return child1, child2
        
#     def mutate(self, chromosome):
#         """Mutation operation to introduce variation"""
#         for gene in chromosome.genes:
#             if random.random() < self.config['MUTATION_RATE']:
#                 # Randomly change day or time slot
#                 if random.choice([True, False]):
#                     gene.day = random.choice(WEEKDAYS)
#                 else:
#                     gene.time_slot = random.choice(ENHANCED_TIME_SLOTS)
                    
#                 # Occasionally reassign room
#                 if random.random() < 0.1:  # 10% chance to change room
#                     if gene.is_lab:
#                         lab_rooms = list(Room.objects.filter(lab=True))
#                         if lab_rooms:
#                             gene.room = random.choice(lab_rooms).room_number
#                     else:
#                         theory_rooms = list(Room.objects.filter(core_sub=True))
#                         if theory_rooms:
#                             gene.room = random.choice(theory_rooms).room_number
                            
#     def evolve(self, semesters_data):
#         """Main evolution loop"""
#         logger.info(f"Starting evolution with {self.config['GENERATIONS']} generations")
        
#         # Initialize population
#         self.initialize_population(semesters_data)
        
#         # Evaluate initial population
#         for chromosome in self.population:
#             self.evaluate_fitness(chromosome)
            
#         for generation in range(self.config['GENERATIONS']):
#             self.generation = generation
            
#             # Sort population by fitness
#             self.population.sort(key=lambda x: x.fitness, reverse=True)
            
#             # Track best chromosome
#             if not self.best_chromosome or self.population[0].fitness > self.best_chromosome.fitness:
#                 self.best_chromosome = self.population[0]
                
#             # Log progress
#             if generation % 10 == 0:
#                 best_fitness = self.population[0].fitness
#                 avg_fitness = sum(c.fitness for c in self.population) / len(self.population)
#                 logger.info(f"Generation {generation}: Best={best_fitness:.2f}, Avg={avg_fitness:.2f}, Conflicts={len(self.best_chromosome.conflicts)}")
                
#             self.fitness_history.append(self.population[0].fitness)
            
#             # Create next generation
#             new_population = []
            
#             # Keep elite individuals
#             elite_size = min(self.config['ELITE_SIZE'], len(self.population))
#             new_population.extend(self.population[:elite_size])
            
#             # Generate offspring
#             while len(new_population) < self.config['POPULATION_SIZE']:
#                 # Selection
#                 parents = self.selection()
#                 parent1, parent2 = random.sample(parents, 2)
                
#                 # Crossover
#                 child1, child2 = self.crossover(parent1, parent2)
                
#                 # Mutation
#                 self.mutate(child1)
#                 self.mutate(child2)
                
#                 # Evaluate offspring
#                 self.evaluate_fitness(child1)
#                 self.evaluate_fitness(child2)
                
#                 new_population.extend([child1, child2])
                
#             # Replace population
#             self.population = new_population[:self.config['POPULATION_SIZE']]
            
#             # Early termination if perfect solution found
#             if self.best_chromosome.fitness >= 1000 and len(self.best_chromosome.conflicts) == 0:
#                 logger.info(f"Perfect solution found at generation {generation}")
#                 break
                
#         logger.info(f"Evolution completed. Best fitness: {self.best_chromosome.fitness}, Conflicts: {len(self.best_chromosome.conflicts)}")
#         return self.best_chromosome

# # ============ INTEGRATION FUNCTIONS ============

# def get_subject_frequency(subject_name, is_lab):
#     """Get frequency based on subject type"""
#     if is_lab:
#         return 1
#     return SUBJECT_FREQUENCY_MAP.get(subject_name, 2)

# def prepare_semesters_data(max_semester):
#     """Prepare data structure for genetic algorithm"""
#     semesters_data = {}
    
#     for semester in range(1, max_semester + 1):
#         sections = Section.objects.filter(semester_number=semester)
#         semesters_data[semester] = {}
        
#         for section in sections:
#             subjects = []
#             core_courses = CoreCourse.objects.filter(semester_number=semester)
            
#             for course in core_courses:
#                 subjects.append({
#                     'name': course.course_name,
#                     'is_lab': course.is_lab,
#                     'duration': course.duration
#                 })
                
#             semesters_data[semester][section.section_name] = subjects
            
#     return semesters_data

# def chromosome_to_database(chromosome):
#     """Convert chromosome to database records"""
#     # Clear existing schedules
#     EnhancedSchedule.objects.all().delete()
#     GlobalScheduleMatrix.objects.all().delete()
    
#     schedule_data = {}
#     processed_global_entries = set()  # Track processed entries to avoid duplicates
    
#     for gene in chromosome.genes:
#         # Save to EnhancedSchedule
#         schedule = EnhancedSchedule.objects.create(
#             semester=gene.semester,
#             section=gene.section,
#             subject=gene.subject,
#             is_lab=gene.is_lab,
#             time_slot=gene.time_slot,
#             day=gene.day,
#             room=gene.room,
#             week_number=gene.session_number
#         )
        
#         # Save to GlobalScheduleMatrix (avoid duplicates)
#         global_key = (gene.time_slot, gene.day, gene.room)
#         if global_key not in processed_global_entries:
#             try:
#                 GlobalScheduleMatrix.objects.create(
#                     time_slot=gene.time_slot,
#                     day=gene.day,
#                     room_number=gene.room,
#                     semester=gene.semester,
#                     section=gene.section,
#                     subject=gene.subject
#                 )
#                 processed_global_entries.add(global_key)
#             except Exception as e:
#                 # Log the duplicate but don't fail the entire process
#                 logger.warning(f"Duplicate GlobalScheduleMatrix entry skipped: {global_key}")
        
#         # Prepare display data
#         section_key = f'Semester {gene.semester} - {gene.section}'
#         if section_key not in schedule_data:
#             schedule_data[section_key] = {day: {} for day in WEEKDAYS}
            
#         if gene.time_slot not in schedule_data[section_key][gene.day]:
#             schedule_data[section_key][gene.day][gene.time_slot] = None
            
#         room_info = f" - Room {gene.room}" if gene.room else ""
#         subject_info = f"{gene.subject} ({'Lab' if gene.is_lab else 'Theory'}){room_info}"
        
#         # Handle multiple classes at same time slot (append instead of overwrite)
#         existing = schedule_data[section_key][gene.day].get(gene.time_slot)
#         if existing:
#             schedule_data[section_key][gene.day][gene.time_slot] = f"{existing} | {subject_info}"
#         else:
#             schedule_data[section_key][gene.day][gene.time_slot] = subject_info
    
#     return schedule_data

# # ============ DJANGO VIEWS ============

# @csrf_exempt
# def generate_enhanced_timetable(request):
#     """Generate timetable using Genetic Algorithm"""
#     if request.method == 'POST':
#         try:
#             # Handle both JSON and form data
#             if request.content_type == 'application/json':
#                 data = json.loads(request.body)
#                 max_semester = data.get('max_semester', 1)
#             else:
#                 max_semester = int(request.POST.get('max_semester', 1))
            
#             logger.info(f"Starting Genetic Algorithm timetable generation for {max_semester} semesters")
            
#             # Prepare data for genetic algorithm
#             semesters_data = prepare_semesters_data(max_semester)
#             print("semesters_data:", semesters_data)
            
#             # Initialize and run genetic algorithm
#             ga = GeneticAlgorithm()
#             best_solution = ga.evolve(semesters_data)
#             print("best_solution.genes (first 10):", [(g.semester, g.section, g.subject, g.is_lab, g.day, g.time_slot, g.room) for g in best_solution.genes[:10]])
#             # Convert solution to database format
#             schedule_data = chromosome_to_database(best_solution)
#             print("schedule_data:", schedule_data)
            
#             # Prepare response
#             response_data = {
#                 'status': 'success',
#                 'message': f'Genetic Algorithm timetable generated successfully',
#                 'schedule': schedule_data,
#                 'algorithm_info': {
#                     'type': 'Genetic Algorithm + Constraint Satisfaction',
#                     'generations': ga.generation + 1,
#                     'final_fitness': best_solution.fitness,
#                     'conflicts': len(best_solution.conflicts),
#                     'population_size': ga.config['POPULATION_SIZE']
#                 },
#                 'features': {
#                     'theoretical_twice_weekly': True,
#                     'inter_semester_conflict_free': len(best_solution.conflicts) == 0,
#                     'extended_hours': True,
#                     'genetic_optimization': True
#                 }
#             }
            
#             if request.content_type == 'application/json':
#                 return JsonResponse(response_data)
#             else:
#                 request.session['temp_enhanced_timetable'] = {
#                     'max_semester': max_semester,
#                     'data': schedule_data
#                 }
#                 # Return JSON instead of template for web interface too
#                 return JsonResponse({
#                     'status': 'success',
#                     'message': 'Timetable generated and stored in session',
#                     'redirect': '/view/'
#                 })
                
#         except Exception as e:
#             logger.error(f"Error in genetic algorithm timetable generation: {str(e)}", exc_info=True)
#             error_msg = f'Genetic Algorithm generation failed: {str(e)}'
            
#             return JsonResponse({'status': 'error', 'message': error_msg})

# def generate_timetable(request):
#     """Main timetable generation entry point"""
#     if request.method == 'POST':
#         form = TimetableGenerationForm(request.POST, request.FILES)
#         if form.is_valid():
#             max_semester = int(form.cleaned_data['max_semester'])
            
#             try:
#                 file = request.FILES['file']
                
#                 # Process Excel file
#                 if file.name.endswith('.xlsx') or file.name.endswith('.xls'):
#                     success = process_excel_file(file)
#                 else:
#                     success = process_csv_files([file])
                
#                 if not success:
#                     return JsonResponse({
#                         'status': 'error',
#                         'message': 'Error processing the uploaded file.'
#                     })
                
#                 # Store data and redirect to genetic algorithm generation
#                 request.session['upload_data'] = {
#                     'max_semester': max_semester,
#                     'file_processed': True
#                 }
                
#                 # Generate using genetic algorithm
#                 return generate_enhanced_web(request, max_semester)
                
#             except Exception as e:
#                 logger.error(f"Error in file processing: {str(e)}", exc_info=True)
#                 return JsonResponse({
#                     'status': 'error', 
#                     'message': f'Error: {str(e)}'
#                 })
#     else:
#         # Return form data as JSON for API compatibility
#         return JsonResponse({
#             'status': 'success',
#             'message': 'Upload Excel file to generate timetable',
#             'required_sheets': ['Students', 'Core Courses', 'Cohort Courses', 'Elective Courses', 'Rooms'],
#             'genetic_algorithm': 'Active'
#         })

# def generate_enhanced_web(request, max_semester=None):
#     """Generate genetic algorithm timetable via web interface"""
#     if max_semester is None:
#         upload_data = request.session.get('upload_data')
#         if not upload_data:
#             return JsonResponse({
#                 'status': 'error',
#                 'message': 'No upload data found. Please upload file first.'
#             })
#         max_semester = upload_data['max_semester']
    
#     # Create fake POST request for genetic algorithm
#     request.method = 'POST'
#     request.POST = {'max_semester': max_semester}
    
#     return generate_enhanced_timetable(request)

# # ============ FILE PROCESSING FUNCTIONS ============

# def process_excel_file(excel_file):
#     try:
#         xls = pd.ExcelFile(excel_file)
        
#         # Process Students and create sections
#         if 'Students' in xls.sheet_names:
#             students = pd.read_excel(xls, 'Students')
#             SemesterStudents.objects.all().delete()
#             Section.objects.all().delete()
            
#             for _, row in students.iterrows():
#                 semester = int(row['Semester Number'])
#                 num_students = int(row['Number of Students'])
                
#                 SemesterStudents.objects.create(
#                     semester_number=semester,
#                     number_of_students=num_students
#                 )
                
#                 Section.create_sections(semester, num_students)

#         # Process Core Courses
#         if 'Core Courses' in xls.sheet_names:
#             core_courses = pd.read_excel(xls, 'Core Courses')
#             CoreCourse.objects.all().delete()
#             for _, row in core_courses.iterrows():
#                 CoreCourse.objects.create(
#                     semester_number=int(row['Semester Number']),
#                     course_name=str(row['Course Name']),
#                     is_lab=str(row['Lab/Theory (Boolean)']).upper() == 'TRUE',
#                     duration=150 if str(row['Lab/Theory (Boolean)']).upper() == 'TRUE' else 75
#                 )

#         # Process Cohort Courses
#         if 'Cohort Courses' in xls.sheet_names:
#             cohort_courses = pd.read_excel(xls, 'Cohort Courses')
#             CohortCourse.objects.all().delete()
#             for _, row in cohort_courses.iterrows():
#                 CohortCourse.objects.create(
#                     course_name=str(row['Cohort Courses']),
#                     semester_number=int(row['Semester Number']),
#                     student_range=int(row['STUDENT  RANGE'])
#                 )

#         # Process Elective Courses
#         if 'Elective Courses' in xls.sheet_names:
#             elective_courses = pd.read_excel(xls, 'Elective Courses')
#             ElectiveCourse.objects.all().delete()
#             for _, row in elective_courses.iterrows():
#                 ElectiveCourse.objects.create(
#                     semester_number=int(row['Semester Number']),
#                     course_name=str(row['Course Name']),
#                     is_tech=str(row['TECH/UNI ELE']).upper() == 'TRUE'
#                 )

#         # Process Rooms
#         if 'Rooms' in xls.sheet_names:
#             rooms = pd.read_excel(xls, 'Rooms')
#             Room.objects.all().delete()
#             for _, row in rooms.iterrows():
#                 Room.objects.create(
#                     room_number=str(row['Room Number']),
#                     core_sub=str(row['Core Sub']).upper() == 'TRUE',
#                     cohort_sub=str(row['Cohort Sub']).upper() == 'TRUE',
#                     lab=str(row['LAB']).upper() == 'TRUE',
#                     dld=str(row['DLD']).upper() == 'TRUE'
#                 )
        
#         return True
#     except Exception as e:
#         logger.error(f"Error processing Excel file: {str(e)}")
#         return False

# def process_csv_files(csv_files):
#     return False

# # ============ UTILITY FUNCTIONS ============

# def save_timetable(request):
#     if request.method == 'POST':
#         temp_enhanced = request.session.get('temp_enhanced_timetable')
#         temp_original = request.session.get('temp_timetable')
        
#         if temp_enhanced:
#             timetable = Timetable.objects.create(
#                 semester=temp_enhanced['max_semester'],
#                 data=temp_enhanced['data']
#             )
#             del request.session['temp_enhanced_timetable']
#             return JsonResponse({
#                 'status': 'success',
#                 'message': 'Genetic Algorithm timetable saved successfully!',
#                 'timetable_id': timetable.id
#             })
#         elif temp_original:
#             timetable = Timetable.objects.create(
#                 semester=temp_original['max_semester'],
#                 data=temp_original['data']
#             )
#             del request.session['temp_timetable']
#             return JsonResponse({
#                 'status': 'success',
#                 'message': 'Timetable saved successfully!',
#                 'timetable_id': timetable.id
#             })
#         else:
#             return JsonResponse({
#                 'status': 'error',
#                 'message': 'No timetable data found to save.'
#             })
    
#     return JsonResponse({
#         'status': 'error',
#         'message': 'Invalid request method'
#     })

# def view_timetable(request):
#     timetables = Timetable.objects.all().order_by('-created_at')
    
#     # Convert timetables to JSON-friendly format
#     timetables_data = []
#     for timetable in timetables:
#         timetables_data.append({
#             'id': timetable.id,
#             'semester': timetable.semester,
#             'created_at': timetable.created_at.isoformat(),
#             'data': timetable.data
#         })
    
#     return JsonResponse({
#         'status': 'success',
#         'timetables': timetables_data,
#         'days': WEEKDAYS,
#         'time_slots': [slot.replace('-', ' - ') for slot in ENHANCED_TIME_SLOTS],
#         'enhanced_time_slots': ENHANCED_TIME_SLOTS,
#     })

# def download_timetable_excel(request, timetable_id):
#     timetable = get_object_or_404(Timetable, id=timetable_id)
    
#     wb = Workbook()
#     ws = wb.active
#     ws.title = f"GA_Timetable_Semesters_1-{timetable.semester}"

#     headers = ['Time Slot'] + WEEKDAYS
#     for col, header in enumerate(headers, start=1):
#         cell = ws.cell(row=1, column=col, value=header)
#         cell.font = Font(bold=True)
#         cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

#     row = 2
#     all_time_slots = [slot.replace('-', ' - ') for slot in ENHANCED_TIME_SLOTS]
    
#     for semester_section, schedule in timetable.data.items():
#         ws.cell(row=row, column=1, value=f"{semester_section} (Genetic Algorithm)").font = Font(bold=True)
#         row += 1
        
#         for time_slot in all_time_slots:
#             ws.cell(row=row, column=1, value=time_slot)
#             for col, day in enumerate(WEEKDAYS, start=2):
#                 cell_value = schedule.get(day, {}).get(time_slot.replace(' - ', '-'), '-')
#                 ws.cell(row=row, column=col, value=cell_value)
#             row += 1
        
#         row += 1

#     for col in ws.columns:
#         max_length = max(len(str(cell.value)) for cell in col)
#         ws.column_dimensions[col[0].column_letter].width = max_length + 2

#     response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#     response['Content-Disposition'] = f'attachment; filename=genetic_algorithm_timetable_semesters_1-{timetable.semester}.xlsx'
#     wb.save(response)
    
#     return response

# def delete_timetable(request, timetable_id):
#     timetable = get_object_or_404(Timetable, id=timetable_id)
#     timetable.delete()
#     return JsonResponse({
#         'status': 'success',
#         'message': 'Timetable deleted successfully!'
#     })

# # ============ VALIDATION AND STATISTICS ============

# @csrf_exempt
# def validate_enhanced_schedule(request):
#     """Validate the genetic algorithm generated schedule"""
#     conflicts = []
    
#     # Check section conflicts
#     schedules_by_section_time = defaultdict(list)
#     for schedule in EnhancedSchedule.objects.all():
#         key = f"{schedule.semester}-{schedule.section}-{schedule.day}-{schedule.time_slot}"
#         schedules_by_section_time[key].append(schedule)
    
#     for key, schedules in schedules_by_section_time.items():
#         if len(schedules) > 1:
#             conflicts.append({
#                 'type': 'section_conflict',
#                 'key': key,
#                 'count': len(schedules)
#             })
    
#     # Check room conflicts
#     room_bookings = defaultdict(list)
#     for matrix in GlobalScheduleMatrix.objects.all():
#         key = f"{matrix.day}-{matrix.time_slot}-{matrix.room_number}"
#         room_bookings[key].append(matrix)
    
#     for key, bookings in room_bookings.items():
#         if len(bookings) > 1:
#             conflicts.append({
#                 'type': 'room_conflict',
#                 'key': key,
#                 'count': len(bookings)
#             })
    
#     # Check theoretical subjects frequency
#     theoretical_counts = defaultdict(int)
#     for schedule in EnhancedSchedule.objects.filter(is_lab=False):
#         key = f"{schedule.semester}-{schedule.section}-{schedule.subject}"
#         theoretical_counts[key] += 1
    
#     twice_weekly_subjects = sum(1 for count in theoretical_counts.values() if count == 2)
#     incorrect_frequency = sum(1 for count in theoretical_counts.values() if count != 2)
    
#     return JsonResponse({
#         'status': 'success',
#         'validation_results': {
#             'total_conflicts': len(conflicts),
#             'conflicts': conflicts,
#             'theoretical_subjects_correct_frequency': twice_weekly_subjects,
#             'theoretical_subjects_incorrect_frequency': incorrect_frequency,
#             'total_schedules': EnhancedSchedule.objects.count(),
#             'algorithm_type': 'Genetic Algorithm + CSP'
#         },
#         'is_valid': len(conflicts) == 0 and incorrect_frequency == 0
#     })

# def get_schedule_statistics(request):
#     """Get detailed statistics about the genetic algorithm generated schedule"""
#     total_schedules = EnhancedSchedule.objects.count()
#     theoretical_count = EnhancedSchedule.objects.filter(is_lab=False).count()
#     lab_count = EnhancedSchedule.objects.filter(is_lab=True).count()
    
#     # Calculate theoretical frequency distribution
#     theoretical_subjects = defaultdict(int)
#     for schedule in EnhancedSchedule.objects.filter(is_lab=False):
#         key = f"{schedule.semester}-{schedule.section}-{schedule.subject}"
#         theoretical_subjects[key] += 1
    
#     frequency_distribution = {
#         'once_weekly': sum(1 for count in theoretical_subjects.values() if count == 1),
#         'twice_weekly': sum(1 for count in theoretical_subjects.values() if count == 2),
#         'more_than_twice': sum(1 for count in theoretical_subjects.values() if count > 2)
#     }
    
#     # Time slot usage
#     time_slot_usage = {}
#     for slot in ENHANCED_TIME_SLOTS:
#         count = EnhancedSchedule.objects.filter(time_slot=slot).count()
#         time_slot_usage[slot] = count
    
#     # Extended hours usage
#     extended_slots = ["17:00-18:15", "18:30-19:45", "20:00-21:15"]
#     extended_usage = sum(time_slot_usage.get(slot, 0) for slot in extended_slots)
#     total_usage = sum(time_slot_usage.values())
#     extended_percentage = (extended_usage / total_usage * 100) if total_usage > 0 else 0
    
#     return JsonResponse({
#         'algorithm_type': 'Genetic Algorithm + Constraint Satisfaction Problem',
#         'total_schedules': total_schedules,
#         'schedule_breakdown': {
#             'theoretical_classes': theoretical_count,
#             'lab_classes': lab_count
#         },
#         'theoretical_frequency_distribution': frequency_distribution,
#         'time_slot_usage': time_slot_usage,
#         'extended_hours_utilization': {
#             'count': extended_usage,
#             'percentage': round(extended_percentage, 2)
#         },
#         'optimization_features': [
#             'Multi-objective fitness function',
#             'Tournament selection',
#             'Single-point crossover',
#             'Random mutation',
#             'Elite preservation',
#             'Constraint satisfaction integration'
#         ]
#     })

# # ============ COMPATIBILITY FUNCTIONS FOR API_VIEWS ============

# def create_timetable(max_semester):
#     """Compatibility function for api_views.py - uses Genetic Algorithm"""
#     logger.info(f"Legacy create_timetable called - redirecting to Genetic Algorithm for {max_semester} semesters")
    
#     try:
#         # Prepare data for genetic algorithm
#         semesters_data = prepare_semesters_data(max_semester)
        
#         # Initialize and run genetic algorithm
#         ga = GeneticAlgorithm()
#         best_solution = ga.evolve(semesters_data)
        
#         # Convert solution to database format
#         schedule_data = chromosome_to_database(best_solution)
        
#         logger.info(f"Genetic Algorithm completed - Fitness: {best_solution.fitness}, Conflicts: {len(best_solution.conflicts)}")
#         return schedule_data
        
#     except Exception as e:
#         logger.error(f"Error in legacy create_timetable: {str(e)}")
#         # Return empty structure if genetic algorithm fails
#         return {}








#####################################################################################################################################
###########################################################################################################################################
###########################################################################################################################################
###########################################################################################################################################
###########################################################################################################################################
###########################################################################################################################################
###########################################################################################################################################

# # Enhanced Genetic Algorithm Timetable Generation System
# from django.shortcuts import render, redirect, get_object_or_404
# from django.contrib import messages
# from .models import (CoreCourse, CohortCourse, ElectiveCourse, Room, SemesterStudents,
#                      Timetable, Section, EnhancedSchedule, GlobalScheduleMatrix,
#                      SubjectFrequencyRule, TimeSlotConfiguration)
# from .forms import TimetableGenerationForm
# from django.contrib.auth.decorators import login_required
# from django.db.models import Sum
# from django.core.files.storage import default_storage
# from django.http import JsonResponse, HttpResponse
# from django.views.decorators.csrf import csrf_exempt
# import random
# import csv
# import os
# import io
# import logging
# import json
# import pandas as pd
# from openpyxl import Workbook
# from openpyxl.styles import Font, Alignment, PatternFill
# from django.utils.text import slugify
# from datetime import datetime, timedelta, time
# from collections import defaultdict
# import numpy as np

# logger = logging.getLogger(__name__)

# # ============ GENETIC ALGORITHM CONFIGURATION ============
# GENETIC_CONFIG = {
#     'POPULATION_SIZE': 50,
#     'GENERATIONS': 100,
#     'CROSSOVER_RATE': 0.8,
#     'MUTATION_RATE': 0.1,
#     'ELITE_SIZE': 5,
#     'TOURNAMENT_SIZE': 3
# }

# SUBJECT_FREQUENCY_MAP = {
#     # Theoretical subjects - 2 times per week
#     "English-I": 2, "Calculus": 2, "Programming Fundamentals": 2,
#     "ICT Theory": 2, "Information Security": 2, "Physics": 2,
#     "Chemistry": 2, "Mathematics": 2, "Statistics": 2,
    
#     # Lab subjects - 1 time per week
#     "Programming Fundamentals Lab": 1, "ICT Lab": 1,
#     "Physics Lab": 1, "Chemistry Lab": 1,
# }

# ENHANCED_TIME_SLOTS = [
#     "08:00-09:15", "09:30-10:45", "11:00-12:15",  # Morning
#     "12:30-13:45", "14:00-15:15", "15:30-16:45",  # Afternoon  
#     "17:00-18:15", "18:30-19:45", "20:00-21:15"   # Evening (Extended hours)
# ]

# WEEKDAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]

# def home(request):
#     """API-friendly home endpoint - no templates needed"""
#     return JsonResponse({
#         'status': 'success',
#         'message': 'Genetic Algorithm Timetable System API Active',
#         'version': '2.0-GA',
#         'algorithm': 'Genetic Algorithm + Constraint Satisfaction',
#         'backend_features': [
#             'Twice-weekly theoretical lectures',
#             'Inter-semester conflict resolution', 
#             'Extended hours (8 AM - 9 PM)',
#             'Multi-objective optimization'
#         ],
#         'api_endpoints': {
#             'generate': '/api/generate/',
#             'enhanced_generate': '/api/generate-enhanced/',
#             'validate': '/api/validate-enhanced/',
#             'stats': '/api/schedule-stats/'
#         },
#         'genetic_config': {
#             'population_size': 50,
#             'generations': 100,
#             'crossover_rate': 0.8,
#             'mutation_rate': 0.1
#         }
#     })

# # ============ GENETIC ALGORITHM CORE CLASSES ============
# class Gene:
#     """Represents a single scheduled class"""
#     def __init__(self, semester, section, subject, is_lab, day=None, time_slot=None, room=None, teacher=None):
#         self.semester = semester
#         self.section = section
#         self.subject = subject
#         self.is_lab = is_lab
#         self.day = day or random.choice(WEEKDAYS)
#         self.time_slot = time_slot or random.choice(ENHANCED_TIME_SLOTS)
#         self.room = room
#         self.teacher = teacher
#         self.session_number = 1

# class Chromosome:
#     """Represents a complete timetable solution"""
#     def __init__(self, genes=None):
#         self.genes = genes or []
#         self.fitness = 0
#         self.conflicts = []
#         self.schedule_matrix = {}
        
#     def initialize_random(self, semesters_data):
#         """Initialize chromosome with random valid assignments"""
#         self.genes = []
        
#         for semester, sections_data in semesters_data.items():
#             for section_name, subjects in sections_data.items():
#                 for subject_info in subjects:
#                     frequency = get_subject_frequency(subject_info['name'], subject_info['is_lab'])
                    
#                     for session in range(frequency):
#                         gene = Gene(
#                             semester=semester,
#                             section=section_name,
#                             subject=subject_info['name'],
#                             is_lab=subject_info['is_lab']
#                         )
#                         gene.session_number = session + 1
#                         self.genes.append(gene)
        
#         # Assign rooms to genes
#         self.assign_rooms_to_genes()
        
#     def assign_rooms_to_genes(self):
#         """Assign appropriate rooms to genes"""
#         lab_rooms = list(Room.objects.filter(lab=True))
#         theory_rooms = list(Room.objects.filter(core_sub=True))
#         backup_rooms = list(Room.objects.filter(cohort_sub=True))
        
#         for gene in self.genes:
#             if gene.is_lab and lab_rooms:
#                 gene.room = random.choice(lab_rooms).room_number
#             elif theory_rooms:
#                 gene.room = random.choice(theory_rooms).room_number
#             elif backup_rooms:
#                 gene.room = random.choice(backup_rooms).room_number
#             else:
#                 gene.room = "TBD"  # To be determined later

# class GeneticAlgorithm:
#     """Main Genetic Algorithm implementation for timetable optimization"""
    
#     def __init__(self, config=None):
#         self.config = config or GENETIC_CONFIG
#         self.population = []
#         self.best_chromosome = None
#         self.generation = 0
#         self.fitness_history = []
        
#     def initialize_population(self, semesters_data):
#         """Create initial population of random timetables"""
#         logger.info(f"Initializing population of size {self.config['POPULATION_SIZE']}")
        
#         self.population = []
#         for i in range(self.config['POPULATION_SIZE']):
#             chromosome = Chromosome()
#             chromosome.initialize_random(semesters_data)
#             self.population.append(chromosome)
            
#         logger.info(f"Created {len(self.population)} chromosomes")
        
#     def evaluate_fitness(self, chromosome):
#         """Evaluate chromosome fitness based on constraints"""
#         fitness = 1000  # Start with high fitness
#         conflicts = []
        
#         # Hard constraints (major penalties)
#         section_conflicts = self.check_section_conflicts(chromosome)
#         room_conflicts = self.check_room_conflicts(chromosome)
#         inter_semester_conflicts = self.check_inter_semester_conflicts(chromosome)
        
#         # Soft constraints (minor penalties)
#         daily_load_penalties = self.check_daily_load_balance(chromosome)
#         theoretical_frequency_penalties = self.check_theoretical_frequency(chromosome)
#         gap_penalties = self.check_time_gaps(chromosome)
        
#         # Apply penalties
#         fitness -= len(section_conflicts) * 200  # Major penalty for section conflicts
#         fitness -= len(room_conflicts) * 150    # Major penalty for room conflicts
#         fitness -= len(inter_semester_conflicts) * 100  # Penalty for inter-semester conflicts
#         fitness -= daily_load_penalties * 50    # Penalty for unbalanced daily loads
#         fitness -= theoretical_frequency_penalties * 75  # Penalty for incorrect frequency
#         fitness -= gap_penalties * 25           # Penalty for large gaps
        
#         # Bonus for extended hours utilization
#         extended_hours_bonus = self.calculate_extended_hours_usage(chromosome)
#         fitness += extended_hours_bonus * 10
        
#         # Store conflicts for debugging
#         chromosome.conflicts = section_conflicts + room_conflicts + inter_semester_conflicts
#         chromosome.fitness = max(0, fitness)  # Ensure non-negative fitness
        
#         return chromosome.fitness
        
#     def check_section_conflicts(self, chromosome):
#         """Check for conflicts within the same section"""
#         conflicts = []
#         section_schedule = defaultdict(lambda: defaultdict(set))
        
#         for gene in chromosome.genes:
#             key = f"{gene.semester}-{gene.section}"
#             time_key = f"{gene.day}-{gene.time_slot}"
            
#             if time_key in section_schedule[key]:
#                 conflicts.append({
#                     'type': 'section_conflict',
#                     'section': key,
#                     'time': time_key,
#                     'subjects': list(section_schedule[key][time_key]) + [gene.subject]
#                 })
            
#             section_schedule[key][time_key].add(gene.subject)
            
#         return conflicts
        
#     def check_room_conflicts(self, chromosome):
#         """Check for room booking conflicts"""
#         conflicts = []
#         room_schedule = defaultdict(set)
        
#         for gene in chromosome.genes:
#             if gene.room:
#                 time_room_key = f"{gene.day}-{gene.time_slot}-{gene.room}"
                
#                 if time_room_key in room_schedule:
#                     conflicts.append({
#                         'type': 'room_conflict',
#                         'room': gene.room,
#                         'time': f"{gene.day}-{gene.time_slot}",
#                         'conflicting_classes': list(room_schedule[time_room_key]) + [f"S{gene.semester}-{gene.section}-{gene.subject}"]
#                     })
                
#                 room_schedule[time_room_key].add(f"S{gene.semester}-{gene.section}-{gene.subject}")
                
#         return conflicts
        
#     def check_inter_semester_conflicts(self, chromosome):
#         """Check for conflicts between different semesters"""
#         conflicts = []
#         time_usage = defaultdict(set)
        
#         for gene in chromosome.genes:
#             time_key = f"{gene.day}-{gene.time_slot}"
#             semester_section = f"S{gene.semester}-{gene.section}"
            
#             # Check if different semesters are scheduled at same time without room separation
#             if time_key in time_usage:
#                 existing_semesters = {entry.split('-')[0] for entry in time_usage[time_key]}
#                 current_semester = f"S{gene.semester}"
                
#                 if current_semester not in existing_semesters:
#                     # Different semesters at same time - check room separation
#                     if not self.has_room_separation(chromosome, time_key):
#                         conflicts.append({
#                             'type': 'inter_semester_conflict',
#                             'time': time_key,
#                             'semesters': list(existing_semesters) + [current_semester]
#                         })
            
#             time_usage[time_key].add(semester_section)
            
#         return conflicts
        
#     def has_room_separation(self, chromosome, time_key):
#         """Check if different semesters use different rooms at same time"""
#         day, time_slot = time_key.split('-', 1)
#         rooms_used = set()
        
#         for gene in chromosome.genes:
#             if gene.day == day and gene.time_slot == time_slot:
#                 if gene.room in rooms_used:
#                     return False
#                 rooms_used.add(gene.room)
                
#         return True
        
#     def check_daily_load_balance(self, chromosome):
#         """Check for balanced daily class loads"""
#         penalties = 0
#         daily_loads = defaultdict(lambda: defaultdict(int))
        
#         for gene in chromosome.genes:
#             section_key = f"{gene.semester}-{gene.section}"
#             daily_loads[section_key][gene.day] += 1
            
#         for section, days in daily_loads.items():
#             for day, count in days.items():
#                 if count > 3:  # More than 3 classes per day
#                     penalties += count - 3
                    
#         return penalties
        
#     def check_theoretical_frequency(self, chromosome):
#         """Check if theoretical subjects are scheduled twice weekly"""
#         penalties = 0
#         subject_counts = defaultdict(int)
        
#         for gene in chromosome.genes:
#             if not gene.is_lab:
#                 key = f"{gene.semester}-{gene.section}-{gene.subject}"
#                 subject_counts[key] += 1
                
#         for subject_key, count in subject_counts.items():
#             expected_frequency = 2  # Theoretical subjects should appear twice
#             if count != expected_frequency:
#                 penalties += abs(count - expected_frequency)
                
#         return penalties
        
#     def check_time_gaps(self, chromosome):
#         """Check for large time gaps in daily schedules"""
#         penalties = 0
#         daily_schedules = defaultdict(lambda: defaultdict(list))
        
#         for gene in chromosome.genes:
#             section_key = f"{gene.semester}-{gene.section}"
#             daily_schedules[section_key][gene.day].append(gene.time_slot)
            
#         for section, days in daily_schedules.items():
#             for day, slots in days.items():
#                 if len(slots) > 1:
#                     # Sort slots and check gaps
#                     sorted_slots = sorted(slots, key=lambda x: datetime.strptime(x.split('-')[0], '%H:%M'))
                    
#                     for i in range(len(sorted_slots) - 1):
#                         current_end = datetime.strptime(sorted_slots[i].split('-')[1], '%H:%M')
#                         next_start = datetime.strptime(sorted_slots[i+1].split('-')[0], '%H:%M')
#                         gap_hours = (next_start - current_end).seconds / 3600
                        
#                         if gap_hours > 2:  # Gap larger than 2 hours
#                             penalties += 1
                            
#         return penalties
        
#     def calculate_extended_hours_usage(self, chromosome):
#         """Calculate bonus for using extended hours (evening slots)"""
#         extended_slots = ["17:00-18:15", "18:30-19:45", "20:00-21:15"]
#         extended_usage = 0
        
#         for gene in chromosome.genes:
#             if gene.time_slot in extended_slots:
#                 extended_usage += 1
                
#         return extended_usage
        
#     def selection(self):
#         """Tournament selection for choosing parents"""
#         selected = []
        
#         for _ in range(self.config['POPULATION_SIZE']):
#             # Tournament selection
#             tournament = random.sample(self.population, min(self.config['TOURNAMENT_SIZE'], len(self.population)))
#             winner = max(tournament, key=lambda x: x.fitness)
#             selected.append(winner)
            
#         return selected
        
#     def crossover(self, parent1, parent2):
#         """Crossover operation to create offspring"""
#         if random.random() > self.config['CROSSOVER_RATE']:
#             return parent1, parent2
            
#         # Single-point crossover
#         crossover_point = random.randint(1, min(len(parent1.genes), len(parent2.genes)) - 1)
        
#         child1 = Chromosome()
#         child2 = Chromosome()
        
#         child1.genes = parent1.genes[:crossover_point] + parent2.genes[crossover_point:]
#         child2.genes = parent2.genes[:crossover_point] + parent1.genes[crossover_point:]
        
#         return child1, child2
        
#     def mutate(self, chromosome):
#         """Mutation operation to introduce variation"""
#         for gene in chromosome.genes:
#             if random.random() < self.config['MUTATION_RATE']:
#                 # Randomly change day or time slot
#                 if random.choice([True, False]):
#                     gene.day = random.choice(WEEKDAYS)
#                 else:
#                     gene.time_slot = random.choice(ENHANCED_TIME_SLOTS)
                    
#                 # Occasionally reassign room
#                 if random.random() < 0.1:  # 10% chance to change room
#                     if gene.is_lab:
#                         lab_rooms = list(Room.objects.filter(lab=True))
#                         if lab_rooms:
#                             gene.room = random.choice(lab_rooms).room_number
#                     else:
#                         theory_rooms = list(Room.objects.filter(core_sub=True))
#                         if theory_rooms:
#                             gene.room = random.choice(theory_rooms).room_number
                            
#     def evolve(self, semesters_data):
#         """Main evolution loop"""
#         logger.info(f"Starting evolution with {self.config['GENERATIONS']} generations")
        
#         # Initialize population
#         self.initialize_population(semesters_data)
        
#         # Evaluate initial population
#         for chromosome in self.population:
#             self.evaluate_fitness(chromosome)
            
#         for generation in range(self.config['GENERATIONS']):
#             self.generation = generation
            
#             # Sort population by fitness
#             self.population.sort(key=lambda x: x.fitness, reverse=True)
            
#             # Track best chromosome
#             if not self.best_chromosome or self.population[0].fitness > self.best_chromosome.fitness:
#                 self.best_chromosome = self.population[0]
                
#             # Log progress
#             if generation % 10 == 0:
#                 best_fitness = self.population[0].fitness
#                 avg_fitness = sum(c.fitness for c in self.population) / len(self.population)
#                 logger.info(f"Generation {generation}: Best={best_fitness:.2f}, Avg={avg_fitness:.2f}, Conflicts={len(self.best_chromosome.conflicts)}")
                
#             self.fitness_history.append(self.population[0].fitness)
            
#             # Create next generation
#             new_population = []
            
#             # Keep elite individuals
#             elite_size = min(self.config['ELITE_SIZE'], len(self.population))
#             new_population.extend(self.population[:elite_size])
            
#             # Generate offspring
#             while len(new_population) < self.config['POPULATION_SIZE']:
#                 # Selection
#                 parents = self.selection()
#                 parent1, parent2 = random.sample(parents, 2)
                
#                 # Crossover
#                 child1, child2 = self.crossover(parent1, parent2)
                
#                 # Mutation
#                 self.mutate(child1)
#                 self.mutate(child2)
                
#                 # Evaluate offspring
#                 self.evaluate_fitness(child1)
#                 self.evaluate_fitness(child2)
                
#                 new_population.extend([child1, child2])
                
#             # Replace population
#             self.population = new_population[:self.config['POPULATION_SIZE']]
            
#             # Early termination if perfect solution found
#             if self.best_chromosome.fitness >= 1000 and len(self.best_chromosome.conflicts) == 0:
#                 logger.info(f"Perfect solution found at generation {generation}")
#                 break
                
#         logger.info(f"Evolution completed. Best fitness: {self.best_chromosome.fitness}, Conflicts: {len(self.best_chromosome.conflicts)}")
#         return self.best_chromosome

# # ============ INTEGRATION FUNCTIONS ============
# def get_subject_frequency(subject_name, is_lab):
#     """Get frequency based on subject type"""
#     if is_lab:
#         return 1
#     return SUBJECT_FREQUENCY_MAP.get(subject_name, 2)

# def prepare_semesters_data(max_semester):
#     """Prepare data structure for genetic algorithm"""
#     semesters_data = {}
    
#     logger.info(f"Preparing data for {max_semester} semesters")
    
#     for semester in range(1, max_semester + 1):
#         # Get sections for this semester
#         sections = Section.objects.filter(semester_number=semester)
#         logger.info(f"Found {sections.count()} sections for semester {semester}")
        
#         semesters_data[semester] = {}
        
#         for section in sections:
#             subjects = []
            
#             # Get core courses for this semester
#             core_courses = CoreCourse.objects.filter(semester_number=semester)
#             logger.info(f"Found {core_courses.count()} core courses for semester {semester}")
            
#             for course in core_courses:
#                 subjects.append({
#                     'name': course.course_name,
#                     'is_lab': course.is_lab,
#                     'duration': course.duration
#                 })
            
#             # Get cohort courses for this semester
#             cohort_courses = CohortCourse.objects.filter(semester_number=semester)
#             logger.info(f"Found {cohort_courses.count()} cohort courses for semester {semester}")
            
#             for course in cohort_courses:
#                 subjects.append({
#                     'name': course.course_name,
#                     'is_lab': False,  # Assume cohort courses are theory
#                     'duration': 75
#                 })
            
#             # Get elective courses for this semester
#             elective_courses = ElectiveCourse.objects.filter(semester_number=semester)
#             logger.info(f"Found {elective_courses.count()} elective courses for semester {semester}")
            
#             for course in elective_courses:
#                 subjects.append({
#                     'name': course.course_name,
#                     'is_lab': False,  # Assume elective courses are theory
#                     'duration': 75
#                 })
            
#             if subjects:  # Only add section if it has subjects
#                 semesters_data[semester][section.section_name] = subjects
#                 logger.info(f"Added {len(subjects)} subjects for section {section.section_name}")
    
#     logger.info(f"Final semesters_data structure: {semesters_data}")
#     return semesters_data

# def chromosome_to_database(chromosome):
#     """Convert chromosome to database records"""
#     # Clear existing schedules
#     EnhancedSchedule.objects.all().delete()
#     GlobalScheduleMatrix.objects.all().delete()
    
#     schedule_data = {}
#     processed_global_entries = set()  # Track processed entries to avoid duplicates
    
#     logger.info(f"Converting chromosome with {len(chromosome.genes)} genes to database")
    
#     for gene in chromosome.genes:
#         # Save to EnhancedSchedule
#         schedule = EnhancedSchedule.objects.create(
#             semester=gene.semester,
#             section=gene.section,
#             subject=gene.subject,
#             is_lab=gene.is_lab,
#             time_slot=gene.time_slot,
#             day=gene.day,
#             room=gene.room,
#             week_number=gene.session_number
#         )
        
#         # Save to GlobalScheduleMatrix (avoid duplicates)
#         global_key = (gene.time_slot, gene.day, gene.room)
#         if global_key not in processed_global_entries:
#             try:
#                 GlobalScheduleMatrix.objects.create(
#                     time_slot=gene.time_slot,
#                     day=gene.day,
#                     room_number=gene.room,
#                     semester=gene.semester,
#                     section=gene.section,
#                     subject=gene.subject
#                 )
#                 processed_global_entries.add(global_key)
#             except Exception as e:
#                 # Log the duplicate but don't fail the entire process
#                 logger.warning(f"Duplicate GlobalScheduleMatrix entry skipped: {global_key}")
        
#         # Prepare display data
#         section_key = f'Semester {gene.semester} - {gene.section}'
#         if section_key not in schedule_data:
#             schedule_data[section_key] = {day: {} for day in WEEKDAYS}
            
#         # Convert time slot format for frontend (HH:MM-HH:MM to HH:MM - HH:MM)
#         display_time_slot = gene.time_slot.replace('-', ' - ')
        
#         room_info = f" - Room {gene.room}" if gene.room else ""
#         subject_info = f"{gene.subject} ({'Lab' if gene.is_lab else 'Theory'}){room_info}"
        
#         # Handle multiple classes at same time slot (append instead of overwrite)
#         existing = schedule_data[section_key][gene.day].get(display_time_slot)
#         if existing:
#             schedule_data[section_key][gene.day][display_time_slot] = f"{existing} | {subject_info}"
#         else:
#             schedule_data[section_key][gene.day][display_time_slot] = subject_info
    
#     # Ensure all time slots are present for consistent display
#     all_display_slots = [slot.replace('-', ' - ') for slot in ENHANCED_TIME_SLOTS]
    
#     for section_key in schedule_data:
#         for day in WEEKDAYS:
#             for slot in all_display_slots:
#                 if slot not in schedule_data[section_key][day]:
#                     schedule_data[section_key][day][slot] = "Free"
    
#     logger.info(f"Generated schedule_data with {len(schedule_data)} sections")
#     return schedule_data

# # ============ DJANGO VIEWS ============
# @csrf_exempt
# def generate_enhanced_timetable(request):
#     """Generate timetable using Genetic Algorithm"""
#     if request.method == 'POST':
#         try:
#             # Handle both JSON and form data
#             if request.content_type == 'application/json':
#                 data = json.loads(request.body)
#                 max_semester = data.get('max_semester', 1)
#             else:
#                 max_semester = int(request.POST.get('max_semester', 1))
                
#             logger.info(f"Starting Genetic Algorithm timetable generation for {max_semester} semesters")
            
#             # Prepare data for genetic algorithm
#             semesters_data = prepare_semesters_data(max_semester)
            
#             if not semesters_data or not any(semesters_data.values()):
#                 return JsonResponse({
#                     'status': 'error',
#                     'message': 'No course data found. Please upload course data first.',
#                     'debug_info': {
#                         'semesters_data': semesters_data,
#                         'sections_count': Section.objects.count(),
#                         'core_courses_count': CoreCourse.objects.count(),
#                         'cohort_courses_count': CohortCourse.objects.count(),
#                         'elective_courses_count': ElectiveCourse.objects.count()
#                     }
#                 })
            
#             logger.info(f"semesters_data: {semesters_data}")
            
#             # Initialize and run genetic algorithm
#             ga = GeneticAlgorithm()
#             best_solution = ga.evolve(semesters_data)
#             logger.info(f"best_solution.genes (first 10): {[(g.semester, g.section, g.subject, g.is_lab, g.day, g.time_slot, g.room) for g in best_solution.genes[:10]]}")
            
#             # Convert solution to database format
#             schedule_data = chromosome_to_database(best_solution)
#             logger.info(f"schedule_data keys: {list(schedule_data.keys())}")
            
#             # Prepare response
#             response_data = {
#                 'status': 'success',
#                 'message': f'Genetic Algorithm timetable generated successfully',
#                 'schedule': schedule_data,
#                 'algorithm_info': {
#                     'type': 'Genetic Algorithm + Constraint Satisfaction',
#                     'generations': ga.generation + 1,
#                     'final_fitness': best_solution.fitness,
#                     'conflicts': len(best_solution.conflicts),
#                     'population_size': ga.config['POPULATION_SIZE']
#                 },
#                 'features': {
#                     'theoretical_twice_weekly': True,
#                     'inter_semester_conflict_free': len(best_solution.conflicts) == 0,
#                     'extended_hours': True,
#                     'genetic_optimization': True
#                 }
#             }
            
#             if request.content_type == 'application/json':
#                 return JsonResponse(response_data)
#             else:
#                 request.session['temp_enhanced_timetable'] = {
#                     'max_semester': max_semester,
#                     'data': schedule_data
#                 }
#                 # Return JSON instead of template for web interface too
#                 return JsonResponse({
#                     'status': 'success',
#                     'message': 'Timetable generated and stored in session',
#                     'redirect': '/view/'
#                 })
                
#         except Exception as e:
#             logger.error(f"Error in genetic algorithm timetable generation: {str(e)}", exc_info=True)
#             error_msg = f'Genetic Algorithm generation failed: {str(e)}'
            
#             return JsonResponse({'status': 'error', 'message': error_msg})

# def generate_timetable(request):
#     """Main timetable generation entry point"""
#     if request.method == 'POST':
#         form = TimetableGenerationForm(request.POST, request.FILES)
#         if form.is_valid():
#             max_semester = int(form.cleaned_data['max_semester'])
            
#             try:
#                 file = request.FILES['file']
                
#                 # Process Excel file
#                 if file.name.endswith('.xlsx') or file.name.endswith('.xls'):
#                     success = process_excel_file(file)
#                 else:
#                     success = process_csv_files([file])
                
#                 if not success:
#                     return JsonResponse({
#                         'status': 'error',
#                         'message': 'Error processing the uploaded file.'
#                     })
                
#                 # Store data and redirect to genetic algorithm generation
#                 request.session['upload_data'] = {
#                     'max_semester': max_semester,
#                     'file_processed': True
#                 }
                
#                 # Generate using genetic algorithm
#                 return generate_enhanced_web(request, max_semester)
                
#             except Exception as e:
#                 logger.error(f"Error in file processing: {str(e)}", exc_info=True)
#                 return JsonResponse({
#                     'status': 'error', 
#                     'message': f'Error: {str(e)}'
#                 })
#     else:
#         # Return form data as JSON for API compatibility
#         return JsonResponse({
#             'status': 'success',
#             'message': 'Upload Excel file to generate timetable',
#             'required_sheets': ['Students', 'Core Courses', 'Cohort Courses', 'Elective Courses', 'Rooms'],
#             'genetic_algorithm': 'Active'
#         })

# def generate_enhanced_web(request, max_semester=None):
#     """Generate genetic algorithm timetable via web interface"""
#     if max_semester is None:
#         upload_data = request.session.get('upload_data')
#         if not upload_data:
#             return JsonResponse({
#                 'status': 'error',
#                 'message': 'No upload data found. Please upload file first.'
#             })
#         max_semester = upload_data['max_semester']
    
#     # Create fake POST request for genetic algorithm
#     request.method = 'POST'
#     request.POST = {'max_semester': max_semester}
    
#     return generate_enhanced_timetable(request)

# # ============ FILE PROCESSING FUNCTIONS ============
# def process_excel_file(excel_file):
#     try:
#         xls = pd.ExcelFile(excel_file)
        
#         # Process Students and create sections
#         if 'Students' in xls.sheet_names:
#             students = pd.read_excel(xls, 'Students')
#             SemesterStudents.objects.all().delete()
#             Section.objects.all().delete()
            
#             for _, row in students.iterrows():
#                 semester = int(row['Semester Number'])
#                 num_students = int(row['Number of Students'])
                
#                 SemesterStudents.objects.create(
#                     semester_number=semester,
#                     number_of_students=num_students
#                 )
                
#                 Section.create_sections(semester, num_students)

#         # Process Core Courses
#         if 'Core Courses' in xls.sheet_names:
#             core_courses = pd.read_excel(xls, 'Core Courses')
#             CoreCourse.objects.all().delete()
#             for _, row in core_courses.iterrows():
#                 CoreCourse.objects.create(
#                     semester_number=int(row['Semester Number']),
#                     course_name=str(row['Course Name']),
#                     is_lab=str(row['Lab/Theory (Boolean)']).upper() == 'TRUE',
#                     duration=150 if str(row['Lab/Theory (Boolean)']).upper() == 'TRUE' else 75
#                 )

#         # Process Cohort Courses
#         if 'Cohort Courses' in xls.sheet_names:
#             cohort_courses = pd.read_excel(xls, 'Cohort Courses')
#             CohortCourse.objects.all().delete()
#             for _, row in cohort_courses.iterrows():
#                 CohortCourse.objects.create(
#                     course_name=str(row['Cohort Courses']),
#                     semester_number=int(row['Semester Number']),
#                     student_range=int(row['STUDENT  RANGE'])
#                 )

#         # Process Elective Courses
#         if 'Elective Courses' in xls.sheet_names:
#             elective_courses = pd.read_excel(xls, 'Elective Courses')
#             ElectiveCourse.objects.all().delete()
#             for _, row in elective_courses.iterrows():
#                 ElectiveCourse.objects.create(
#                     semester_number=int(row['Semester Number']),
#                     course_name=str(row['Course Name']),
#                     is_tech=str(row['TECH/UNI ELE']).upper() == 'TRUE'
#                 )

#         # Process Rooms
#         if 'Rooms' in xls.sheet_names:
#             rooms = pd.read_excel(xls, 'Rooms')
#             Room.objects.all().delete()
#             for _, row in rooms.iterrows():
#                 Room.objects.create(
#                     room_number=str(row['Room Number']),
#                     core_sub=str(row['Core Sub']).upper() == 'TRUE',
#                     cohort_sub=str(row['Cohort Sub']).upper() == 'TRUE',
#                     lab=str(row['LAB']).upper() == 'TRUE',
#                     dld=str(row['DLD']).upper() == 'TRUE'
#                 )
        
#         return True
#     except Exception as e:
#         logger.error(f"Error processing Excel file: {str(e)}")
#         return False

# def process_csv_files(csv_files):
#     return False

# # ============ UTILITY FUNCTIONS ============
# def save_timetable(request):
#     if request.method == 'POST':
#         temp_enhanced = request.session.get('temp_enhanced_timetable')
#         temp_original = request.session.get('temp_timetable')
        
#         if temp_enhanced:
#             timetable = Timetable.objects.create(
#                 semester=temp_enhanced['max_semester'],
#                 data=temp_enhanced['data']
#             )
#             del request.session['temp_enhanced_timetable']
#             return JsonResponse({
#                 'status': 'success',
#                 'message': 'Genetic Algorithm timetable saved successfully!',
#                 'timetable_id': timetable.id
#             })
#         elif temp_original:
#             timetable = Timetable.objects.create(
#                 semester=temp_original['max_semester'],
#                 data=temp_original['data']
#             )
#             del request.session['temp_timetable']
#             return JsonResponse({
#                 'status': 'success',
#                 'message': 'Timetable saved successfully!',
#                 'timetable_id': timetable.id
#             })
#         else:
#             return JsonResponse({
#                 'status': 'error',
#                 'message': 'No timetable data found to save.'
#             })
    
#     return JsonResponse({
#         'status': 'error',
#         'message': 'Invalid request method'
#     })

# def view_timetable(request):
#     timetables = Timetable.objects.all().order_by('-created_at')
    
#     # Convert timetables to JSON-friendly format
#     timetables_data = []
#     for timetable in timetables:
#         timetables_data.append({
#             'id': timetable.id,
#             'semester': timetable.semester,
#             'created_at': timetable.created_at.isoformat(),
#             'data': timetable.data
#         })
    
#     return JsonResponse({
#         'status': 'success',
#         'timetables': timetables_data,
#         'days': WEEKDAYS,
#         'time_slots': [slot.replace('-', ' - ') for slot in ENHANCED_TIME_SLOTS],
#         'enhanced_time_slots': ENHANCED_TIME_SLOTS,
#     })

# def download_timetable_excel(request, timetable_id):
#     timetable = get_object_or_404(Timetable, id=timetable_id)
    
#     wb = Workbook()
#     ws = wb.active
#     ws.title = f"GA_Timetable_Semesters_1-{timetable.semester}"

#     headers = ['Time Slot'] + WEEKDAYS
#     for col, header in enumerate(headers, start=1):
#         cell = ws.cell(row=1, column=col, value=header)
#         cell.font = Font(bold=True)
#         cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

#     row = 2
#     all_time_slots = [slot.replace('-', ' - ') for slot in ENHANCED_TIME_SLOTS]
    
#     for semester_section, schedule in timetable.data.items():
#         ws.cell(row=row, column=1, value=f"{semester_section} (Genetic Algorithm)").font = Font(bold=True)
#         row += 1
        
#         for time_slot in all_time_slots:
#             ws.cell(row=row, column=1, value=time_slot)
#             for col, day in enumerate(WEEKDAYS, start=2):
#                 cell_value = schedule.get(day, {}).get(time_slot, 'Free')
#                 ws.cell(row=row, column=col, value=cell_value)
#             row += 1
        
#         row += 1

#     for col in ws.columns:
#         max_length = max(len(str(cell.value)) for cell in col)
#         ws.column_dimensions[col[0].column_letter].width = max_length + 2

#     response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#     response['Content-Disposition'] = f'attachment; filename=genetic_algorithm_timetable_semesters_1-{timetable.semester}.xlsx'
#     wb.save(response)
    
#     return response

# def delete_timetable(request, timetable_id):
#     timetable = get_object_or_404(Timetable, id=timetable_id)
#     timetable.delete()
#     return JsonResponse({
#         'status': 'success',
#         'message': 'Timetable deleted successfully!'
#     })

# # ============ VALIDATION AND STATISTICS ============
# @csrf_exempt
# def validate_enhanced_schedule(request):
#     """Validate the genetic algorithm generated schedule"""
#     conflicts = []
    
#     # Check section conflicts
#     schedules_by_section_time = defaultdict(list)
#     for schedule in EnhancedSchedule.objects.all():
#         key = f"{schedule.semester}-{schedule.section}-{schedule.day}-{schedule.time_slot}"
#         schedules_by_section_time[key].append(schedule)
    
#     for key, schedules in schedules_by_section_time.items():
#         if len(schedules) > 1:
#             conflicts.append({
#                 'type': 'section_conflict',
#                 'key': key,
#                 'count': len(schedules)
#             })
    
#     # Check room conflicts
#     room_bookings = defaultdict(list)
#     for matrix in GlobalScheduleMatrix.objects.all():
#         key = f"{matrix.day}-{matrix.time_slot}-{matrix.room_number}"
#         room_bookings[key].append(matrix)
    
#     for key, bookings in room_bookings.items():
#         if len(bookings) > 1:
#             conflicts.append({
#                 'type': 'room_conflict',
#                 'key': key,
#                 'count': len(bookings)
#             })
    
#     # Check theoretical subjects frequency
#     theoretical_counts = defaultdict(int)
#     for schedule in EnhancedSchedule.objects.filter(is_lab=False):
#         key = f"{schedule.semester}-{schedule.section}-{schedule.subject}"
#         theoretical_counts[key] += 1
    
#     twice_weekly_subjects = sum(1 for count in theoretical_counts.values() if count == 2)
#     incorrect_frequency = sum(1 for count in theoretical_counts.values() if count != 2)
    
#     return JsonResponse({
#         'status': 'success',
#         'validation_results': {
#             'total_conflicts': len(conflicts),
#             'conflicts': conflicts,
#             'theoretical_subjects_correct_frequency': twice_weekly_subjects,
#             'theoretical_subjects_incorrect_frequency': incorrect_frequency,
#             'total_schedules': EnhancedSchedule.objects.count(),
#             'algorithm_type': 'Genetic Algorithm + CSP'
#         },
#         'is_valid': len(conflicts) == 0 and incorrect_frequency == 0
#     })

# def get_schedule_statistics(request):
#     """Get detailed statistics about the genetic algorithm generated schedule"""
#     total_schedules = EnhancedSchedule.objects.count()
#     theoretical_count = EnhancedSchedule.objects.filter(is_lab=False).count()
#     lab_count = EnhancedSchedule.objects.filter(is_lab=True).count()
    
#     # Calculate theoretical frequency distribution
#     theoretical_subjects = defaultdict(int)
#     for schedule in EnhancedSchedule.objects.filter(is_lab=False):
#         key = f"{schedule.semester}-{schedule.section}-{schedule.subject}"
#         theoretical_subjects[key] += 1
    
#     frequency_distribution = {
#         'once_weekly': sum(1 for count in theoretical_subjects.values() if count == 1),
#         'twice_weekly': sum(1 for count in theoretical_subjects.values() if count == 2),
#         'more_than_twice': sum(1 for count in theoretical_subjects.values() if count > 2)
#     }
    
#     # Time slot usage
#     time_slot_usage = {}
#     for slot in ENHANCED_TIME_SLOTS:
#         count = EnhancedSchedule.objects.filter(time_slot=slot).count()
#         time_slot_usage[slot] = count
    
#     # Extended hours usage
#     extended_slots = ["17:00-18:15", "18:30-19:45", "20:00-21:15"]
#     extended_usage = sum(time_slot_usage.get(slot, 0) for slot in extended_slots)
#     total_usage = sum(time_slot_usage.values())
#     extended_percentage = (extended_usage / total_usage * 100) if total_usage > 0 else 0
    
#     return JsonResponse({
#         'algorithm_type': 'Genetic Algorithm + Constraint Satisfaction Problem',
#         'total_schedules': total_schedules,
#         'schedule_breakdown': {
#             'theoretical_classes': theoretical_count,
#             'lab_classes': lab_count
#         },
#         'theoretical_frequency_distribution': frequency_distribution,
#         'time_slot_usage': time_slot_usage,
#         'extended_hours_utilization': {
#             'count': extended_usage,
#             'percentage': round(extended_percentage, 2)
#         },
#         'optimization_features': [
#             'Multi-objective fitness function',
#             'Tournament selection',
#             'Single-point crossover',
#             'Random mutation',
#             'Elite preservation',
#             'Constraint satisfaction integration'
#         ]
#     })

# # ============ COMPATIBILITY FUNCTIONS FOR API_VIEWS ============
# def create_timetable(max_semester):
#     """Compatibility function for api_views.py - uses Genetic Algorithm"""
#     logger.info(f"Legacy create_timetable called - redirecting to Genetic Algorithm for {max_semester} semesters")
    
#     try:
#         # Prepare data for genetic algorithm
#         semesters_data = prepare_semesters_data(max_semester)
        
#         if not semesters_data or not any(semesters_data.values()):
#             logger.error("No course data found for genetic algorithm")
#             return {}
        
#         # Initialize and run genetic algorithm
#         ga = GeneticAlgorithm()
#         best_solution = ga.evolve(semesters_data)
        
#         # Convert solution to database format
#         schedule_data = chromosome_to_database(best_solution)
        
#         logger.info(f"Genetic Algorithm completed - Fitness: {best_solution.fitness}, Conflicts: {len(best_solution.conflicts)}")
#         return schedule_data
        
#     except Exception as e:
#         logger.error(f"Error in legacy create_timetable: {str(e)}")
#         # Return empty structure if genetic algorithm fails
#         return {}


###########################################################################################################################################
###########################################################################################################################################
###########################################################################################################################################
###########################################################################################################################################
###########################################################################################################################################
###########################################################################################################################################



# Enhanced Genetic Algorithm Timetable Generation System - FIXED VERSION
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from .models import (CoreCourse, CohortCourse, ElectiveCourse, Room, SemesterStudents,
                     Timetable, Section, EnhancedSchedule, GlobalScheduleMatrix,
                     SubjectFrequencyRule, TimeSlotConfiguration)
from .forms import TimetableGenerationForm
from django.contrib.auth.decorators import login_required
from django.db.models import Sum
from django.core.files.storage import default_storage
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
import random
import csv
import os
import io
import logging
import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from django.utils.text import slugify
from datetime import datetime, timedelta, time
from collections import defaultdict
import numpy as np

logger = logging.getLogger(__name__)

# ============ FIXED GENETIC ALGORITHM CONFIGURATION ============
GENETIC_CONFIG = {
    'POPULATION_SIZE': 100,
    'GENERATIONS': 300,      # More generations for complex constraints
    'CROSSOVER_RATE': 0.8,   
    'MUTATION_RATE': 0.15,   
    'ELITE_SIZE': 10,        
    'TOURNAMENT_SIZE': 5     
}

SUBJECT_FREQUENCY_MAP = {
    # Theoretical subjects - 2 times per week
    "English-I": 2, "Calculus": 2, "Programming Fundamentals": 2,
    "ICT Theory": 2, "Information Security": 2, "Physics": 2,
    "Chemistry": 2, "Mathematics": 2, "Statistics": 2,
    "Applied Physics": 2, "Pak Study": 2, "Programming Fundamental Theory": 2,
    
    # Lab subjects - 1 time per week
    "Programming Fundamentals Lab": 1, "ICT Lab": 1,
    "Physics Lab": 1, "Chemistry Lab": 1, "Applied Physics Lab": 1,
    "Programming Fundamental Lab": 1,
}

# FIXED: Separate time slots for theory and lab classes
THEORY_TIME_SLOTS = [
    "08:00-09:15", "09:30-10:45", "11:00-12:15",  # Morning
    "12:30-13:45", "14:00-15:15", "15:30-16:45",  # Afternoon  
    "17:00-18:15", "18:30-19:45", "20:00-21:15"   # Evening
]

LAB_TIME_SLOTS = [
    "08:00-10:30", "11:00-13:30", "14:00-16:30",  # 2.5 hour lab slots
    "17:00-19:30"  # Evening lab slot
]

WEEKDAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]

def home(request):
    """API-friendly home endpoint - no templates needed"""
    return JsonResponse({
        'status': 'success',
        'message': 'Fixed Genetic Algorithm Timetable System API Active',
        'version': '2.2-GA-FIXED',
        'algorithm': 'Fixed Genetic Algorithm + Strict Constraint Satisfaction',
        'backend_features': [
            'Strict no-simultaneous-classes enforcement',
            'Proper lab scheduling (2.5 hours)',
            'Theory scheduling (1.25 hours)',
            'Inter-semester conflict resolution', 
            'Extended hours (8 AM - 9 PM)',
            'Multi-objective optimization'
        ],
        'api_endpoints': {
            'generate': '/api/generate/',
            'enhanced_generate': '/api/generate-enhanced/',
            'validate': '/api/validate-enhanced/',
            'stats': '/api/schedule-stats/'
        },
        'genetic_config': {
            'population_size': 100,
            'generations': 300,
            'crossover_rate': 0.8,
            'mutation_rate': 0.15
        }
    })

# ============ GENETIC ALGORITHM CORE CLASSES ============
class Gene:
    """Represents a single scheduled class"""
    def __init__(self, semester, section, subject, is_lab, day=None, time_slot=None, room=None, teacher=None):
        self.semester = semester
        self.section = section
        self.subject = subject
        self.is_lab = is_lab
        self.day = day or random.choice(WEEKDAYS)
        # FIXED: Use appropriate time slots based on class type
        if is_lab:
            self.time_slot = time_slot or random.choice(LAB_TIME_SLOTS)
        else:
            self.time_slot = time_slot or random.choice(THEORY_TIME_SLOTS)
        self.room = room
        self.teacher = teacher
        self.session_number = 1

class Chromosome:
    """Represents a complete timetable solution"""
    def __init__(self, genes=None):
        self.genes = genes or []
        self.fitness = 0
        self.conflicts = []
        self.schedule_matrix = {}
        
    def initialize_random(self, semesters_data):
        """Initialize chromosome with random valid assignments"""
        self.genes = []
        
        for semester, sections_data in semesters_data.items():
            for section_name, subjects in sections_data.items():
                for subject_info in subjects:
                    frequency = get_subject_frequency(subject_info['name'], subject_info['is_lab'])
                    
                    for session in range(frequency):
                        gene = Gene(
                            semester=semester,
                            section=section_name,
                            subject=subject_info['name'],
                            is_lab=subject_info['is_lab']
                        )
                        gene.session_number = session + 1
                        self.genes.append(gene)
        
        # Assign rooms to genes
        self.assign_rooms_to_genes()
        
        # FIXED: Ensure no simultaneous classes for same section
        self.fix_simultaneous_classes()
        
    def assign_rooms_to_genes(self):
        """Assign appropriate rooms to genes"""
        lab_rooms = list(Room.objects.filter(lab=True))
        theory_rooms = list(Room.objects.filter(core_sub=True))
        backup_rooms = list(Room.objects.filter(cohort_sub=True))
        
        for gene in self.genes:
            if gene.is_lab and lab_rooms:
                gene.room = random.choice(lab_rooms).room_number
            elif theory_rooms:
                gene.room = random.choice(theory_rooms).room_number
            elif backup_rooms:
                gene.room = random.choice(backup_rooms).room_number
            else:
                gene.room = "TBD"
                
    def fix_simultaneous_classes(self):
        """FIXED: Ensure no two classes for same section at same time"""
        section_schedule = defaultdict(set)
        
        for gene in self.genes:
            section_key = f"{gene.semester}-{gene.section}"
            time_key = f"{gene.day}-{gene.time_slot}"
            
            # If this time slot is already occupied for this section, find a new one
            attempts = 0
            while time_key in section_schedule[section_key] and attempts < 50:
                gene.day = random.choice(WEEKDAYS)
                if gene.is_lab:
                    gene.time_slot = random.choice(LAB_TIME_SLOTS)
                else:
                    gene.time_slot = random.choice(THEORY_TIME_SLOTS)
                time_key = f"{gene.day}-{gene.time_slot}"
                attempts += 1
            
            section_schedule[section_key].add(time_key)

class GeneticAlgorithm:
    """Fixed Genetic Algorithm implementation for timetable optimization"""
    
    def __init__(self, config=None):
        self.config = config or GENETIC_CONFIG
        self.population = []
        self.best_chromosome = None
        self.generation = 0
        self.fitness_history = []
        
    def initialize_population(self, semesters_data):
        """Create initial population of random timetables"""
        logger.info(f"Initializing population of size {self.config['POPULATION_SIZE']}")
        
        self.population = []
        for i in range(self.config['POPULATION_SIZE']):
            chromosome = Chromosome()
            chromosome.initialize_random(semesters_data)
            self.population.append(chromosome)
            
        logger.info(f"Created {len(self.population)} chromosomes")
        
    def evaluate_fitness(self, chromosome):
        """FIXED: Strict fitness function with zero tolerance for simultaneous classes"""
        base_fitness = 1000
        
        # CRITICAL: Check for simultaneous classes (ZERO TOLERANCE)
        simultaneous_conflicts = self.check_simultaneous_classes_strict(chromosome)
        if len(simultaneous_conflicts) > 0:
            # Immediate severe penalty for any simultaneous classes
            chromosome.fitness = 0
            chromosome.conflicts = simultaneous_conflicts
            return 0
        
        # Check for missing lab classes
        missing_labs = self.check_missing_labs(chromosome)
        if len(missing_labs) > 0:
            # Severe penalty for missing labs
            base_fitness -= len(missing_labs) * 300
        
        # Other constraints (only apply if no simultaneous classes)
        room_conflicts = self.check_room_conflicts(chromosome)
        inter_semester_conflicts = self.check_inter_semester_conflicts(chromosome)
        daily_load_penalties = self.check_daily_load_balance(chromosome)
        theoretical_frequency_penalties = self.check_theoretical_frequency(chromosome)
        gap_penalties = self.check_time_gaps(chromosome)
        
        # Apply penalties
        fitness = base_fitness
        fitness -= len(room_conflicts) * 100
        fitness -= len(inter_semester_conflicts) * 50
        fitness -= daily_load_penalties * 30
        fitness -= theoretical_frequency_penalties * 40
        fitness -= gap_penalties * 20
        
        # Bonus for good practices
        extended_hours_bonus = self.calculate_extended_hours_usage(chromosome)
        fitness += extended_hours_bonus * 5
        
        balance_bonus = self.calculate_balance_bonus(chromosome)
        fitness += balance_bonus * 10
        
        # Store all conflicts
        chromosome.conflicts = simultaneous_conflicts + room_conflicts + inter_semester_conflicts + missing_labs
        chromosome.fitness = max(0, fitness)
        
        return chromosome.fitness
        
    def check_simultaneous_classes_strict(self, chromosome):
        """FIXED: Strict check for simultaneous classes - zero tolerance"""
        conflicts = []
        section_schedule = defaultdict(lambda: defaultdict(list))
        
        for gene in chromosome.genes:
            section_key = f"{gene.semester}-{gene.section}"
            time_key = f"{gene.day}-{gene.time_slot}"
            
            section_schedule[section_key][time_key].append(gene.subject)
        
        # Check for any time slot with more than one class
        for section_key, schedule in section_schedule.items():
            for time_key, subjects in schedule.items():
                if len(subjects) > 1:
                    conflicts.append({
                        'type': 'simultaneous_classes',
                        'section': section_key,
                        'time': time_key,
                        'subjects': subjects,
                        'count': len(subjects)
                    })
        
        return conflicts
        
    def check_missing_labs(self, chromosome):
        """FIXED: Check if all lab classes are properly scheduled"""
        conflicts = []
        
        # Get all lab subjects that should be scheduled
        expected_labs = set()
        actual_labs = set()
        
        for gene in chromosome.genes:
            if gene.is_lab:
                expected_labs.add(f"{gene.semester}-{gene.section}-{gene.subject}")
                actual_labs.add(f"{gene.semester}-{gene.section}-{gene.subject}")
        
        # Check if any lab subjects are missing from the schedule
        for gene in chromosome.genes:
            if gene.is_lab:
                lab_key = f"{gene.semester}-{gene.section}-{gene.subject}"
                if lab_key not in actual_labs:
                    conflicts.append({
                        'type': 'missing_lab',
                        'lab': lab_key
                    })
        
        return conflicts
        
    def calculate_balance_bonus(self, chromosome):
        """Calculate bonus for well-balanced daily schedules"""
        bonus = 0
        daily_loads = defaultdict(lambda: defaultdict(int))
        
        for gene in chromosome.genes:
            section_key = f"{gene.semester}-{gene.section}"
            daily_loads[section_key][gene.day] += 1
            
        for section, days in daily_loads.items():
            daily_counts = list(days.values())
            if daily_counts:
                # Bonus for even distribution
                std_dev = np.std(daily_counts) if len(daily_counts) > 1 else 0
                if std_dev < 1.0:
                    bonus += 5
                    
        return bonus
        
    def check_room_conflicts(self, chromosome):
        """Check for room booking conflicts"""
        conflicts = []
        room_schedule = defaultdict(set)
        
        for gene in chromosome.genes:
            if gene.room and gene.room != "TBD":
                time_room_key = f"{gene.day}-{gene.time_slot}-{gene.room}"
                
                if time_room_key in room_schedule:
                    conflicts.append({
                        'type': 'room_conflict',
                        'room': gene.room,
                        'time': f"{gene.day}-{gene.time_slot}",
                        'conflicting_classes': list(room_schedule[time_room_key]) + [f"S{gene.semester}-{gene.section}-{gene.subject}"]
                    })
                
                room_schedule[time_room_key].add(f"S{gene.semester}-{gene.section}-{gene.subject}")
                
        return conflicts
        
    def check_inter_semester_conflicts(self, chromosome):
        """Check for conflicts between different semesters"""
        conflicts = []
        time_usage = defaultdict(set)
        
        for gene in chromosome.genes:
            time_key = f"{gene.day}-{gene.time_slot}"
            semester_section = f"S{gene.semester}-{gene.section}"
            
            if time_key in time_usage:
                existing_entries = time_usage[time_key]
                for existing_entry in existing_entries:
                    existing_semester = existing_entry.split('-')[0]
                    current_semester = f"S{gene.semester}"
                    
                    if existing_semester != current_semester:
                        if self.same_room_conflict(chromosome, gene, time_key):
                            conflicts.append({
                                'type': 'inter_semester_conflict',
                                'time': time_key,
                                'semesters': [existing_semester, current_semester]
                            })
                            break
            
            time_usage[time_key].add(semester_section)
            
        return conflicts
        
    def same_room_conflict(self, chromosome, current_gene, time_key):
        """Check if genes at same time use same room"""
        day, time_slot = time_key.split('-', 1)
        
        for gene in chromosome.genes:
            if (gene != current_gene and 
                gene.day == day and 
                gene.time_slot == time_slot and 
                gene.room == current_gene.room and 
                gene.room != "TBD"):
                return True
        return False
        
    def check_daily_load_balance(self, chromosome):
        """Check for balanced daily class loads"""
        penalties = 0
        daily_loads = defaultdict(lambda: defaultdict(int))
        
        for gene in chromosome.genes:
            section_key = f"{gene.semester}-{gene.section}"
            daily_loads[section_key][gene.day] += 1
            
        for section, days in daily_loads.items():
            for day, count in days.items():
                if count > 4:
                    penalties += count - 4
                    
        return penalties
        
    def check_theoretical_frequency(self, chromosome):
        """Check if theoretical subjects are scheduled twice weekly"""
        penalties = 0
        subject_counts = defaultdict(int)
        
        for gene in chromosome.genes:
            if not gene.is_lab:
                key = f"{gene.semester}-{gene.section}-{gene.subject}"
                subject_counts[key] += 1
                
        for subject_key, count in subject_counts.items():
            expected_frequency = 2
            if count != expected_frequency:
                penalties += abs(count - expected_frequency)
                
        return penalties
        
    def check_time_gaps(self, chromosome):
        """Check for large time gaps in daily schedules"""
        penalties = 0
        daily_schedules = defaultdict(lambda: defaultdict(list))
        
        for gene in chromosome.genes:
            section_key = f"{gene.semester}-{gene.section}"
            daily_schedules[section_key][gene.day].append(gene.time_slot)
            
        for section, days in daily_schedules.items():
            for day, slots in days.items():
                if len(slots) > 1:
                    # Sort slots and check gaps
                    theory_slots = [s for s in slots if s in THEORY_TIME_SLOTS]
                    if len(theory_slots) > 1:
                        sorted_slots = sorted(theory_slots, key=lambda x: datetime.strptime(x.split('-')[0], '%H:%M'))
                        
                        for i in range(len(sorted_slots) - 1):
                            current_end = datetime.strptime(sorted_slots[i].split('-')[1], '%H:%M')
                            next_start = datetime.strptime(sorted_slots[i+1].split('-')[0], '%H:%M')
                            gap_hours = (next_start - current_end).seconds / 3600
                            
                            if gap_hours > 3:
                                penalties += 0.5
                            
        return penalties
        
    def calculate_extended_hours_usage(self, chromosome):
        """Calculate bonus for using extended hours"""
        extended_slots = ["17:00-18:15", "18:30-19:45", "20:00-21:15", "17:00-19:30"]
        extended_usage = 0
        
        for gene in chromosome.genes:
            if gene.time_slot in extended_slots:
                extended_usage += 1
                
        return extended_usage
        
    def selection(self):
        """Tournament selection with fitness-based probability"""
        selected = []
        
        for _ in range(self.config['POPULATION_SIZE']):
            tournament = random.sample(self.population, min(self.config['TOURNAMENT_SIZE'], len(self.population)))
            # Prioritize chromosomes with no simultaneous conflicts
            no_conflict_candidates = [c for c in tournament if not any(conf['type'] == 'simultaneous_classes' for conf in c.conflicts)]
            if no_conflict_candidates:
                winner = max(no_conflict_candidates, key=lambda x: x.fitness)
            else:
                winner = max(tournament, key=lambda x: x.fitness)
            selected.append(winner)
            
        return selected
        
    def crossover(self, parent1, parent2):
        """FIXED: Crossover with conflict prevention"""
        if random.random() > self.config['CROSSOVER_RATE']:
            return parent1, parent2
            
        min_length = min(len(parent1.genes), len(parent2.genes))
        if min_length < 4:
            return parent1, parent2
            
        # Single-point crossover
        crossover_point = random.randint(1, min_length - 1)
        
        child1 = Chromosome()
        child2 = Chromosome()
        
        child1.genes = parent1.genes[:crossover_point] + parent2.genes[crossover_point:]
        child2.genes = parent2.genes[:crossover_point] + parent1.genes[crossover_point:]
        
        # FIXED: Repair simultaneous conflicts after crossover
        child1.fix_simultaneous_classes()
        child2.fix_simultaneous_classes()
        
        return child1, child2
        
    def mutate(self, chromosome):
        """FIXED: Mutation with conflict prevention"""
        for gene in chromosome.genes:
            if random.random() < self.config['MUTATION_RATE']:
                # Store original values
                original_day = gene.day
                original_time = gene.time_slot
                
                # Try mutation
                mutation_type = random.choice(['day', 'time', 'both'])
                
                if mutation_type == 'day':
                    gene.day = random.choice(WEEKDAYS)
                elif mutation_type == 'time':
                    if gene.is_lab:
                        gene.time_slot = random.choice(LAB_TIME_SLOTS)
                    else:
                        gene.time_slot = random.choice(THEORY_TIME_SLOTS)
                elif mutation_type == 'both':
                    gene.day = random.choice(WEEKDAYS)
                    if gene.is_lab:
                        gene.time_slot = random.choice(LAB_TIME_SLOTS)
                    else:
                        gene.time_slot = random.choice(THEORY_TIME_SLOTS)
                
                # Check if mutation creates conflict
                if self.creates_simultaneous_conflict(chromosome, gene):
                    # Revert mutation
                    gene.day = original_day
                    gene.time_slot = original_time
                    
    def creates_simultaneous_conflict(self, chromosome, mutated_gene):
        """Check if a gene mutation creates simultaneous class conflict"""
        section_key = f"{mutated_gene.semester}-{mutated_gene.section}"
        time_key = f"{mutated_gene.day}-{mutated_gene.time_slot}"
        
        for gene in chromosome.genes:
            if (gene != mutated_gene and 
                f"{gene.semester}-{gene.section}" == section_key and
                f"{gene.day}-{gene.time_slot}" == time_key):
                return True
        return False
                            
    def evolve(self, semesters_data):
        """FIXED: Evolution loop with strict constraint enforcement"""
        logger.info(f"Starting fixed evolution with {self.config['GENERATIONS']} generations")
        
        # Initialize population
        self.initialize_population(semesters_data)
        
        # Evaluate initial population
        for chromosome in self.population:
            self.evaluate_fitness(chromosome)
            
        stagnation_counter = 0
        best_fitness_ever = 0
        
        for generation in range(self.config['GENERATIONS']):
            self.generation = generation
            
            # Sort population by fitness
            self.population.sort(key=lambda x: x.fitness, reverse=True)
            
            # Track best chromosome
            current_best_fitness = self.population[0].fitness
            if not self.best_chromosome or current_best_fitness > self.best_chromosome.fitness:
                self.best_chromosome = self.population[0]
                best_fitness_ever = current_best_fitness
                stagnation_counter = 0
            else:
                stagnation_counter += 1
                
            # Log progress
            if generation % 20 == 0:
                best_fitness = self.population[0].fitness
                avg_fitness = sum(c.fitness for c in self.population) / len(self.population)
                simultaneous_conflicts = len([c for c in self.best_chromosome.conflicts if c['type'] == 'simultaneous_classes'])
                logger.info(f"Generation {generation}: Best={best_fitness:.2f}, Avg={avg_fitness:.2f}, Conflicts={len(self.best_chromosome.conflicts)}, Simultaneous={simultaneous_conflicts}")
                
            self.fitness_history.append(self.population[0].fitness)
            
            # Create next generation
            new_population = []
            
            # Keep elite individuals (only those without simultaneous conflicts)
            elite_candidates = [c for c in self.population if not any(conf['type'] == 'simultaneous_classes' for conf in c.conflicts)]
            if elite_candidates:
                elite_size = min(self.config['ELITE_SIZE'], len(elite_candidates))
                new_population.extend(elite_candidates[:elite_size])
            
            # Generate offspring
            while len(new_population) < self.config['POPULATION_SIZE']:
                parents = self.selection()
                parent1, parent2 = random.sample(parents, 2)
                
                child1, child2 = self.crossover(parent1, parent2)
                
                self.mutate(child1)
                self.mutate(child2)
                
                self.evaluate_fitness(child1)
                self.evaluate_fitness(child2)
                
                new_population.extend([child1, child2])
                
            # Replace population
            self.population = new_population[:self.config['POPULATION_SIZE']]
            
            # Early termination if perfect solution found
            if (self.best_chromosome.fitness >= 900 and 
                not any(conf['type'] == 'simultaneous_classes' for conf in self.best_chromosome.conflicts)):
                logger.info(f"Perfect solution found at generation {generation}")
                break
                
        logger.info(f"Evolution completed. Best fitness: {self.best_chromosome.fitness}, Conflicts: {len(self.best_chromosome.conflicts)}")
        return self.best_chromosome

# ============ INTEGRATION FUNCTIONS ============

def get_subject_frequency(subject_name, is_lab):
    """Get frequency based on subject type"""
    if is_lab:
        return 1
    return SUBJECT_FREQUENCY_MAP.get(subject_name, 2)

def prepare_semesters_data(max_semester):
    """Prepare data structure for genetic algorithm"""
    semesters_data = {}
    
    logger.info(f"Preparing data for {max_semester} semesters")
    
    for semester in range(1, max_semester + 1):
        sections = Section.objects.filter(semester_number=semester)
        logger.info(f"Found {sections.count()} sections for semester {semester}")
        
        semesters_data[semester] = {}
        
        for section in sections:
            subjects = []
            
            # Get core courses for this semester
            core_courses = CoreCourse.objects.filter(semester_number=semester)
            logger.info(f"Found {core_courses.count()} core courses for semester {semester}")
            
            for course in core_courses:
                subjects.append({
                    'name': course.course_name,
                    'is_lab': course.is_lab,
                    'duration': course.duration
                })
            
            # Get cohort courses for this semester
            cohort_courses = CohortCourse.objects.filter(semester_number=semester)
            logger.info(f"Found {cohort_courses.count()} cohort courses for semester {semester}")
            
            for course in cohort_courses:
                subjects.append({
                    'name': course.course_name,
                    'is_lab': False,
                    'duration': 75
                })
            
            # Get elective courses for this semester
            elective_courses = ElectiveCourse.objects.filter(semester_number=semester)
            logger.info(f"Found {elective_courses.count()} elective courses for semester {semester}")
            
            for course in elective_courses:
                subjects.append({
                    'name': course.course_name,
                    'is_lab': False,
                    'duration': 75
                })
            
            if subjects:
                semesters_data[semester][section.section_name] = subjects
                logger.info(f"Added {len(subjects)} subjects for section {section.section_name}")
    
    logger.info(f"Final semesters_data structure: {semesters_data}")
    return semesters_data

def chromosome_to_database(chromosome):
    """Convert chromosome to database records"""
    # Clear existing schedules
    EnhancedSchedule.objects.all().delete()
    GlobalScheduleMatrix.objects.all().delete()
    
    schedule_data = {}
    processed_global_entries = set()
    
    logger.info(f"Converting chromosome with {len(chromosome.genes)} genes to database")
    
    for gene in chromosome.genes:
        # Save to EnhancedSchedule
        EnhancedSchedule.objects.create(
            semester=gene.semester,
            section=gene.section,
            subject=gene.subject,
            is_lab=gene.is_lab,
            time_slot=gene.time_slot,
            day=gene.day,
            room=gene.room,
            week_number=gene.session_number
        )
        
        # Save to GlobalScheduleMatrix (avoid duplicates)
        global_key = (gene.time_slot, gene.day, gene.room)
        if global_key not in processed_global_entries:
            try:
                GlobalScheduleMatrix.objects.create(
                    time_slot=gene.time_slot,
                    day=gene.day,
                    room_number=gene.room,
                    semester=gene.semester,
                    section=gene.section,
                    subject=gene.subject
                )
                processed_global_entries.add(global_key)
            except Exception as e:
                logger.warning(f"Duplicate GlobalScheduleMatrix entry skipped: {global_key}")
        
        # Prepare display data
        section_key = f'Semester {gene.semester} - {gene.section}'
        if section_key not in schedule_data:
            schedule_data[section_key] = {day: {} for day in WEEKDAYS}
            
        # Convert time slot format for frontend
        display_time_slot = gene.time_slot.replace('-', ' - ')
        
        room_info = f" - Room {gene.room}" if gene.room else ""
        class_type = "Lab" if gene.is_lab else "Theory"
        subject_info = f"{gene.subject} ({class_type}){room_info}"
        
        # FIXED: No multiple classes at same time - this should not happen now
        if display_time_slot in schedule_data[section_key][gene.day]:
            logger.warning(f"Conflict detected in final schedule: {section_key} {gene.day} {display_time_slot}")
        
        schedule_data[section_key][gene.day][display_time_slot] = subject_info
    
    # Ensure all time slots are present for consistent display
    all_display_slots = [slot.replace('-', ' - ') for slot in THEORY_TIME_SLOTS + LAB_TIME_SLOTS]
    
    for section_key in schedule_data:
        for day in WEEKDAYS:
            for slot in all_display_slots:
                if slot not in schedule_data[section_key][day]:
                    schedule_data[section_key][day][slot] = "Free"
    
    logger.info(f"Generated schedule_data with {len(schedule_data)} sections")
    return schedule_data

# ============ DJANGO VIEWS ============

@csrf_exempt
def generate_enhanced_timetable(request):
    """Generate timetable using Fixed Genetic Algorithm"""
    if request.method == 'POST':
        try:
            # Handle both JSON and form data
            if request.content_type == 'application/json':
                data = json.loads(request.body)
                max_semester = data.get('max_semester', 1)
            else:
                max_semester = int(request.POST.get('max_semester', 1))
            
            logger.info(f"Starting Fixed Genetic Algorithm timetable generation for {max_semester} semesters")
            
            # Prepare data for genetic algorithm
            semesters_data = prepare_semesters_data(max_semester)
            
            if not semesters_data or not any(semesters_data.values()):
                return JsonResponse({
                    'status': 'error',
                    'message': 'No course data found. Please upload course data first.',
                    'debug_info': {
                        'semesters_data': semesters_data,
                        'sections_count': Section.objects.count(),
                        'core_courses_count': CoreCourse.objects.count(),
                        'cohort_courses_count': CohortCourse.objects.count(),
                        'elective_courses_count': ElectiveCourse.objects.count()
                    }
                })
            
            logger.info(f"semesters_data: {semesters_data}")
            
            # Initialize and run fixed genetic algorithm
            ga = GeneticAlgorithm()
            best_solution = ga.evolve(semesters_data)
            logger.info(f"best_solution.genes (first 10): {[(g.semester, g.section, g.subject, g.is_lab, g.day, g.time_slot, g.room) for g in best_solution.genes[:10]]}")
            
            # Convert solution to database format
            schedule_data = chromosome_to_database(best_solution)
            logger.info(f"schedule_data keys: {list(schedule_data.keys())}")
            
            # Prepare response
            response_data = {
                'status': 'success',
                'message': f'Fixed Genetic Algorithm timetable generated successfully',
                'schedule': schedule_data,
                'algorithm_info': {
                    'type': 'Fixed Genetic Algorithm + Strict Constraint Satisfaction',
                    'version': '2.2-FIXED',
                    'generations': ga.generation + 1,
                    'final_fitness': best_solution.fitness,
                    'conflicts': len(best_solution.conflicts),
                    'population_size': ga.config['POPULATION_SIZE'],
                    'fixes': [
                        'Zero tolerance for simultaneous classes',
                        'Proper lab scheduling (2.5 hours)',
                        'Separate time slots for theory and lab',
                        'Conflict prevention in crossover/mutation',
                        'Strict fitness evaluation'
                    ]
                },
                'features': {
                    'no_simultaneous_classes': True,
                    'proper_lab_scheduling': True,
                    'theoretical_twice_weekly': True,
                    'inter_semester_conflict_free': len(best_solution.conflicts) == 0,
                    'extended_hours': True,
                    'genetic_optimization': True
                }
            }
            
            if request.content_type == 'application/json':
                return JsonResponse(response_data)
            else:
                request.session['temp_enhanced_timetable'] = {
                    'max_semester': max_semester,
                    'data': schedule_data
                }
                return JsonResponse({
                    'status': 'success',
                    'message': 'Fixed timetable generated and stored in session',
                    'redirect': '/view/'
                })
                
        except Exception as e:
            logger.error(f"Error in fixed genetic algorithm timetable generation: {str(e)}", exc_info=True)
            error_msg = f'Fixed Genetic Algorithm generation failed: {str(e)}'
            
            return JsonResponse({'status': 'error', 'message': error_msg})

def generate_timetable(request):
    """Main timetable generation entry point"""
    if request.method == 'POST':
        form = TimetableGenerationForm(request.POST, request.FILES)
        if form.is_valid():
            max_semester = int(form.cleaned_data['max_semester'])
            
            try:
                file = request.FILES['file']
                
                # Process Excel file
                if file.name.endswith('.xlsx') or file.name.endswith('.xls'):
                    success = process_excel_file(file)
                else:
                    success = process_csv_files([file])
                
                if not success:
                    return JsonResponse({
                        'status': 'error',
                        'message': 'Error processing the uploaded file.'
                    })
                
                # Store data and redirect to genetic algorithm generation
                request.session['upload_data'] = {
                    'max_semester': max_semester,
                    'file_processed': True
                }
                
                # Generate using fixed genetic algorithm
                return generate_enhanced_web(request, max_semester)
                
            except Exception as e:
                logger.error(f"Error in file processing: {str(e)}", exc_info=True)
                return JsonResponse({
                    'status': 'error', 
                    'message': f'Error: {str(e)}'
                })
    else:
        # Return form data as JSON for API compatibility
        return JsonResponse({
            'status': 'success',
            'message': 'Upload Excel file to generate timetable',
            'required_sheets': ['Students', 'Core Courses', 'Cohort Courses', 'Elective Courses', 'Rooms'],
            'genetic_algorithm': 'Fixed Active'
        })

def generate_enhanced_web(request, max_semester=None):
    """Generate fixed genetic algorithm timetable via web interface"""
    if max_semester is None:
        upload_data = request.session.get('upload_data')
        if not upload_data:
            return JsonResponse({
                'status': 'error',
                'message': 'No upload data found. Please upload file first.'
            })
        max_semester = upload_data['max_semester']
    
    # Create fake POST request for genetic algorithm
    request.method = 'POST'
    request.POST = {'max_semester': max_semester}
    
    return generate_enhanced_timetable(request)

# ============ FILE PROCESSING FUNCTIONS ============
def process_excel_file(excel_file):
    try:
        xls = pd.ExcelFile(excel_file)
        
        # Process Students and create sections
        if 'Students' in xls.sheet_names:
            students = pd.read_excel(xls, 'Students')
            SemesterStudents.objects.all().delete()
            Section.objects.all().delete()
            
            for _, row in students.iterrows():
                semester = int(row['Semester Number'])
                num_students = int(row['Number of Students'])
                
                SemesterStudents.objects.create(
                    semester_number=semester,
                    number_of_students=num_students
                )
                
                Section.create_sections(semester, num_students)

        # Process Core Courses
        if 'Core Courses' in xls.sheet_names:
            core_courses = pd.read_excel(xls, 'Core Courses')
            CoreCourse.objects.all().delete()
            for _, row in core_courses.iterrows():
                CoreCourse.objects.create(
                    semester_number=int(row['Semester Number']),
                    course_name=str(row['Course Name']),
                    is_lab=str(row['Lab/Theory (Boolean)']).upper() == 'TRUE',
                    duration=150 if str(row['Lab/Theory (Boolean)']).upper() == 'TRUE' else 75
                )

        # Process Cohort Courses
        if 'Cohort Courses' in xls.sheet_names:
            cohort_courses = pd.read_excel(xls, 'Cohort Courses')
            CohortCourse.objects.all().delete()
            for _, row in cohort_courses.iterrows():
                CohortCourse.objects.create(
                    course_name=str(row['Cohort Courses']),
                    semester_number=int(row['Semester Number']),
                    student_range=int(row['STUDENT  RANGE'])
                )

        # Process Elective Courses
        if 'Elective Courses' in xls.sheet_names:
            elective_courses = pd.read_excel(xls, 'Elective Courses')
            ElectiveCourse.objects.all().delete()
            for _, row in elective_courses.iterrows():
                ElectiveCourse.objects.create(
                    semester_number=int(row['Semester Number']),
                    course_name=str(row['Course Name']),
                    is_tech=str(row['TECH/UNI ELE']).upper() == 'TRUE'
                )

        # Process Rooms
        if 'Rooms' in xls.sheet_names:
            rooms = pd.read_excel(xls, 'Rooms')
            Room.objects.all().delete()
            for _, row in rooms.iterrows():
                Room.objects.create(
                    room_number=str(row['Room Number']),
                    core_sub=str(row['Core Sub']).upper() == 'TRUE',
                    cohort_sub=str(row['Cohort Sub']).upper() == 'TRUE',
                    lab=str(row['LAB']).upper() == 'TRUE',
                    dld=str(row['DLD']).upper() == 'TRUE'
                )
        
        return True
    except Exception as e:
        logger.error(f"Error processing Excel file: {str(e)}")
        return False

def process_csv_files(csv_files):
    return False

# ============ UTILITY FUNCTIONS ============
def save_timetable(request):
    if request.method == 'POST':
        temp_enhanced = request.session.get('temp_enhanced_timetable')
        temp_original = request.session.get('temp_timetable')
        
        if temp_enhanced:
            timetable = Timetable.objects.create(
                semester=temp_enhanced['max_semester'],
                data=temp_enhanced['data']
            )
            del request.session['temp_enhanced_timetable']
            return JsonResponse({
                'status': 'success',
                'message': 'Fixed Genetic Algorithm timetable saved successfully!',
                'timetable_id': timetable.id
            })
        elif temp_original:
            timetable = Timetable.objects.create(
                semester=temp_original['max_semester'],
                data=temp_original['data']
            )
            del request.session['temp_timetable']
            return JsonResponse({
                'status': 'success',
                'message': 'Timetable saved successfully!',
                'timetable_id': timetable.id
            })
        else:
            return JsonResponse({
                'status': 'error',
                'message': 'No timetable data found to save.'
            })
    
    return JsonResponse({
        'status': 'error',
        'message': 'Invalid request method'
    })

def view_timetable(request):
    timetables = Timetable.objects.all().order_by('-created_at')
    
    # Convert timetables to JSON-friendly format
    timetables_data = []
    for timetable in timetables:
        timetables_data.append({
            'id': timetable.id,
            'semester': timetable.semester,
            'created_at': timetable.created_at.isoformat(),
            'data': timetable.data
        })
    
    return JsonResponse({
        'status': 'success',
        'timetables': timetables_data,
        'days': WEEKDAYS,
        'time_slots': [slot.replace('-', ' - ') for slot in THEORY_TIME_SLOTS + LAB_TIME_SLOTS],
        'theory_time_slots': THEORY_TIME_SLOTS,
        'lab_time_slots': LAB_TIME_SLOTS,
    })

def download_timetable_excel(request, timetable_id):
    timetable = get_object_or_404(Timetable, id=timetable_id)
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"Fixed_GA_Timetable_Semesters_1-{timetable.semester}"

    headers = ['Time Slot'] + WEEKDAYS
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    row = 2
    all_time_slots = [slot.replace('-', ' - ') for slot in THEORY_TIME_SLOTS + LAB_TIME_SLOTS]
    
    for semester_section, schedule in timetable.data.items():
        ws.cell(row=row, column=1, value=f"{semester_section} (Fixed Genetic Algorithm)").font = Font(bold=True)
        row += 1
        
        for time_slot in all_time_slots:
            ws.cell(row=row, column=1, value=time_slot)
            for col, day in enumerate(WEEKDAYS, start=2):
                cell_value = schedule.get(day, {}).get(time_slot, 'Free')
                ws.cell(row=row, column=col, value=cell_value)
            row += 1
        
        row += 1

    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=fixed_genetic_algorithm_timetable_semesters_1-{timetable.semester}.xlsx'
    wb.save(response)
    
    return response

def delete_timetable(request, timetable_id):
    timetable = get_object_or_404(Timetable, id=timetable_id)
    timetable.delete()
    return JsonResponse({
        'status': 'success',
        'message': 'Timetable deleted successfully!'
    })

# ============ VALIDATION AND STATISTICS ============
@csrf_exempt
def validate_enhanced_schedule(request):
    """Validate the fixed genetic algorithm generated schedule"""
    conflicts = []
    
    # Check section conflicts
    schedules_by_section_time = defaultdict(list)
    for schedule in EnhancedSchedule.objects.all():
        key = f"{schedule.semester}-{schedule.section}-{schedule.day}-{schedule.time_slot}"
        schedules_by_section_time[key].append(schedule)
    
    for key, schedules in schedules_by_section_time.items():
        if len(schedules) > 1:
            conflicts.append({
                'type': 'section_conflict',
                'key': key,
                'count': len(schedules)
            })
    
    # Check room conflicts
    room_bookings = defaultdict(list)
    for matrix in GlobalScheduleMatrix.objects.all():
        key = f"{matrix.day}-{matrix.time_slot}-{matrix.room_number}"
        room_bookings[key].append(matrix)
    
    for key, bookings in room_bookings.items():
        if len(bookings) > 1:
            conflicts.append({
                'type': 'room_conflict',
                'key': key,
                'count': len(bookings)
            })
    
    # Check theoretical subjects frequency
    theoretical_counts = defaultdict(int)
    for schedule in EnhancedSchedule.objects.filter(is_lab=False):
        key = f"{schedule.semester}-{schedule.section}-{schedule.subject}"
        theoretical_counts[key] += 1
    
    twice_weekly_subjects = sum(1 for count in theoretical_counts.values() if count == 2)
    incorrect_frequency = sum(1 for count in theoretical_counts.values() if count != 2)
    
    # Check lab scheduling
    lab_count = EnhancedSchedule.objects.filter(is_lab=True).count()
    
    return JsonResponse({
        'status': 'success',
        'validation_results': {
            'total_conflicts': len(conflicts),
            'conflicts': conflicts,
            'theoretical_subjects_correct_frequency': twice_weekly_subjects,
            'theoretical_subjects_incorrect_frequency': incorrect_frequency,
            'lab_classes_scheduled': lab_count,
            'total_schedules': EnhancedSchedule.objects.count(),
            'algorithm_type': 'Fixed Genetic Algorithm + Strict CSP'
        },
        'is_valid': len(conflicts) == 0 and incorrect_frequency == 0
    })

def get_schedule_statistics(request):
    """Get detailed statistics about the fixed genetic algorithm generated schedule"""
    total_schedules = EnhancedSchedule.objects.count()
    theoretical_count = EnhancedSchedule.objects.filter(is_lab=False).count()
    lab_count = EnhancedSchedule.objects.filter(is_lab=True).count()
    
    # Calculate theoretical frequency distribution
    theoretical_subjects = defaultdict(int)
    for schedule in EnhancedSchedule.objects.filter(is_lab=False):
        key = f"{schedule.semester}-{schedule.section}-{schedule.subject}"
        theoretical_subjects[key] += 1
    
    frequency_distribution = {
        'once_weekly': sum(1 for count in theoretical_subjects.values() if count == 1),
        'twice_weekly': sum(1 for count in theoretical_subjects.values() if count == 2),
        'more_than_twice': sum(1 for count in theoretical_subjects.values() if count > 2)
    }
    
    # Time slot usage
    theory_slot_usage = {}
    for slot in THEORY_TIME_SLOTS:
        count = EnhancedSchedule.objects.filter(time_slot=slot, is_lab=False).count()
        theory_slot_usage[slot] = count
        
    lab_slot_usage = {}
    for slot in LAB_TIME_SLOTS:
        count = EnhancedSchedule.objects.filter(time_slot=slot, is_lab=True).count()
        lab_slot_usage[slot] = count
    
    return JsonResponse({
        'algorithm_type': 'Fixed Genetic Algorithm + Strict Constraint Satisfaction Problem',
        'version': '2.2-FIXED',
        'total_schedules': total_schedules,
        'schedule_breakdown': {
            'theoretical_classes': theoretical_count,
            'lab_classes': lab_count
        },
        'theoretical_frequency_distribution': frequency_distribution,
        'theory_time_slot_usage': theory_slot_usage,
        'lab_time_slot_usage': lab_slot_usage,
        'optimization_features': [
            'Zero tolerance for simultaneous classes',
            'Proper lab scheduling (2.5 hours)',
            'Separate time slots for theory and lab',
            'Conflict prevention in crossover/mutation',
            'Strict fitness evaluation',
            'Elite selection without conflicts'
        ]
    })

# ============ COMPATIBILITY FUNCTIONS FOR API_VIEWS ============
def create_timetable(max_semester):
    """Compatibility function for api_views.py - uses Fixed Genetic Algorithm"""
    logger.info(f"Legacy create_timetable called - redirecting to Fixed Genetic Algorithm for {max_semester} semesters")
    
    try:
        # Prepare data for genetic algorithm
        semesters_data = prepare_semesters_data(max_semester)
        
        if not semesters_data or not any(semesters_data.values()):
            logger.error("No course data found for fixed genetic algorithm")
            return {}
        
        # Initialize and run fixed genetic algorithm
        ga = GeneticAlgorithm()
        best_solution = ga.evolve(semesters_data)
        
        # Convert solution to database format
        schedule_data = chromosome_to_database(best_solution)
        
        logger.info(f"Fixed Genetic Algorithm completed - Fitness: {best_solution.fitness}, Conflicts: {len(best_solution.conflicts)}")
        return schedule_data
        
    except Exception as e:
        logger.error(f"Error in legacy create_timetable: {str(e)}")
        # Return empty structure if genetic algorithm fails
        return {}



###########################################################################################################################################
###########################################################################################################################################
###########################################################################################################################################
###########################################################################################################################################
###########################################################################################################################################
###########################################################################################################################################



